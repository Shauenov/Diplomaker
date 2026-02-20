import openpyxl
import pandas as pd

print("=" * 70)
print("TEST EXCEL STRUCTURE CHECK")
print("=" * 70)

wb = openpyxl.load_workbook('test_grades.xlsx')
ws = wb['3F-1']

print("\nSubjects (Row 2):")
for col in range(3, 23, 4):  # Columns C, G, K, O, S (every 4th starting from 3)
    subject = ws.cell(2, col).value
    if subject:
        print(f"  Col {col}: {subject}")

print("\nStudent 1 (Row 6) - Иванов grades:")
for col in range(3, 23):
    val = ws.cell(6, col).value
    if val is not None and val != "":
        print(f"  Col {col}: {val}")

print("\n" + "=" * 70)
print("PARSED DATA CHECK")
print("=" * 70)

df = pd.read_excel('test_grades.xlsx', sheet_name='3F-1', header=None)
print(f"\nDataFrame shape: {df.shape}")
print(f"Row 5 (student 1): {df.iloc[5].tolist()[:15]}")
