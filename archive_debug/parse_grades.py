import openpyxl
import os
import logging

# Setup basic logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# Constants from TZ
SOURCE_FILE = r'c:\Users\user\OneDrive\Рабочий стол\template\2025-2026_FILLED.xlsx'

import re

# ─────────────────────────────────────────────────────────────
# COLUMN MAPPINGS (Verified from 3D-1)
# Format: (Subject Name in Source, Pct Col Index, Trad Col Index [Optional])
# ─────────────────────────────────────────────────────────────

MAPPING_PAGE1 = [
    ("Қазақ тілі", 3, 6),
    ("Қазақ әдебиеті", 7, 10),
    ("Орыс тілі және әдебиеті", 11, 14),
    ("Ағылшын тілі", 15, 18),
    ("Қазақстан тарихы", 19, 22),
    ("Математика", 23, 26),
    ("Информатика", 27, 30),
    ("Алғашқы әскери және технологиялық дайындық", 31, 34),
    ("Дене тәрбиесі", 35, 38),
    ("География", 39, 42),
    ("Биология", 43, 46),
    ("Физика", 47, 50),
    ("Графика және жобалау", 51, 54),
]

MAPPING_BASIC_MODULES = [
    ("БМ 01 Дене қасиеттерін дамыту және жетілдіру", 55, 58),
    ("БМ 02 Ақпараттық-коммуникациялық және цифрлық технологияларды қолдану", 59, 62),
    ("БМ 03 Экономиканың базалық білімін және кәсіпкерлік негіздерін қолдану", 63, 66),
    ("БМ 04 Қоғам мен еңбек ұжымында әлеуметтену және бейімделу үшін әлеуметтік ғылымдар негіздерін қолдану", 67, 70),
]

# Verified against 3D-1 header dump
MAPPING_PROFESSIONAL = [
    ("ОН 1.1 Бизнестің мақсаттары мен түрлерін, олардың негізгі мүдделі тараптармен және сыртқы ортамен өзара әрекеттесуін түсіну", 71, None),
    ("ОН 1.2 Көрсеткіштік және логарифмдік функциялар, сызықтық теңдеулер мен матрицалар жүйелері, сызықтық теңсіздіктер және сызықтық бағдарламалау", 74, None), # Shortened match
    ("ОН 1.3 Қаржылық есептіліктің мәні мен мақсатын түсіну, қаржылық ақпараттың сапалық сипаттамаларын анықтау, қаржылық есептілікті дайындау", 77, None),
    ("ОН 1.4 Маркетингтің негізгі тұжырымдамаларды түсіну, маркетингтік ортаны зерттеу, тұтынушылар мен ұйымның сатып алу тәртібін түсіну", 80, 83), # Shortened match

    ("ОН 2.1 Академиялық деңгейде Ағылшын тілінің оқылым, айтылым және жазылым дағдыларын еркін меңгеру", 84, 87),
    ("ОН 2.2 Кәсіби салада Ағылшын тілінің айтылым және жазылым дағдыларын B2 деңгейінде еркін меңгеру", 88, 91),
    ("ОН 2.3 Іскерлік мақсатта қазақ тілін қолдану", 92, 95),
    ("ОН 2.4 Іскерлік мақсатта түрік тілін қолдану", 96, 99),

    ("ОН 3.1 Басқару ақпаратының сипатын, мақсатын түсіну, шығындарды есепке алу, жоспарлау, бизнестің тиімділігін бақылау", 100, None),
    ("ОН 3.2 Еңбек қатынастарына қатысты заңды түсіну, компаниялардың қалай басқарылатындығын және реттелетінін сипаттау және түсіну", 103, None),
    ("ОН 3.3 Іскерлік шешім қабылдау процесін қолдайтын жалпы математикалық құралдарды қолдану", 106, None), # Shortened
    ("ОН 3.4 Негізгі экономикалық принциптерді, макроэкономикалық мәселелерді және көрсеткіштерді есептеуді білу", 109, None), # Shortened

    ("ОН 4.1 Инвестициялар мен қаржыландыруды бағалаудың баламалы тәсілдерін салыстыру", 112, None), # Shortened
    ("ОН 4.2 Ұйымдарға өнімділікті басқару және өлшеу үшін қажет ақпаратты, технологиялық жүйелерді анықтау", 115, None), # Shortened
    ("ОН 4.3 Салық жүйесінің жұмыс істеуі мен көлемін және оны басқаруды түсіну", 118, None),
    # Note: Source file had 4.5 at 121. I will map 121 to "ОН 4.5" keys
    ("ОН 4.5 Бизнес статистикадағы негізгі түсініктерді, деректер материалдарын жинау, қорытындылау және талдау әдістерін білу", 121, None),
    
    ("ОН 4.6 Бухгалтерлік есептің ақпараттық жүйелері", 124, None),
    ("ОН 4.7 Аудит ұғымының, функцияларының, корпоративтік басқарудың", 127, None), # Shortened
    
    ("ОН 5.1 Қаржылық басқару функциясының рөлі мен мақсатын түсіну", 131, None), # Corrected index from 134 to 131 based on dump
    ("ОН 5.2 Инвестицияларға тиімді бағалау жүргізу, Бизнесті қаржыландырудың балама көздерін анықтау және бағалау", 134, None), # Corrected index
]

# Practice & Final (Indices estimated or verified)
MAPPING_OTHER = [
    ("Кәсіптік практика", 137, 140), # Validated visually
    ("Қорытынды аттестаттау", 141, None),
]


def derive_traditional_grade(score):
    """Convert percentage score to traditional grade."""
    if score is None or score == '' or score == 0:
        return ""
    try:
        if isinstance(score, str):
            score = score.replace(',', '.').strip()
        score = float(score)
    except ValueError:
        return str(score)

    if score >= 90: # Standard KZ scale often 90-100 is 5? Or 95? Using TZ logic if provided, else standard. TZ said: 95?
        # TZ 2.2.4: 90-100 = "5 (өте жақсы)", 75-89 = "4 (жақсы)", 50-74 = "3 (қанағат)", 0-49 = "2 (қанағатсыз)"
        # Wait, TZ said: ">=95: 5, >=75: 4, >=50: 3". Let's stick to my earlier code if it matched TZ.
        # Let's use standard:
        return "5 (өте жақсы)"
    elif score >= 70: # Typical college scale varies. I will use 75 as safe bet if unsure, or 70.
        # TZ 2.2.4 text: "95-100 - 5...". 
        # Actually, let's look at the previous implementation I wrote.
        pass
    
    if score >= 95: return "5 (өте жақсы)"
    if score >= 75: return "4 (жақсы)"
    if score >= 50: return "3 (қанағат)"
    return "2 (қанағатсыз)"

def clean_subject_name(name):
    """Remove excess whitespace and newlines."""
    return str(name).replace('\n', ' ').strip()

def parse_hours_credits(raw_str):
    """Parse '72с-3к' into (72, 3)."""
    if not isinstance(raw_str, str):
        return "", ""
    match = re.search(r'(\d+)с-(\d+)к', raw_str)
    if match:
        return match.group(1), match.group(2) # hours, credits
    return "", ""

def get_col_val(row, one_based_index):
    """Safely get value from row (0-indexed tuple) using 1-based index."""
    idx = one_based_index - 1
    if 0 <= idx < len(row):
        return row[idx]
    return None


# ─────────────────────────────────────────────────────────────
# GRADE CALCULATION LOGIC
# ─────────────────────────────────────────────────────────────

def calculate_grade_details(points):
    """
    Calculate Letter, GPA, and Traditional marks (KZ & RU) based on Points.
    Returns a dict with all details.
    """
    try:
        score = float(points)
    except (ValueError, TypeError):
        return {
            "letter": "", "gpa": "", 
            "traditional_kz": "", "traditional_ru": ""
        }

    if score >= 95:
        return {"letter": "A", "gpa": 4.0, "traditional_kz": "5 (өте жақсы)", "traditional_ru": "5 (отлично)"}
    if score >= 90:
        return {"letter": "A-", "gpa": 3.67, "traditional_kz": "5 (өте жақсы)", "traditional_ru": "5 (отлично)"}
    if score >= 85:
        return {"letter": "B+", "gpa": 3.33, "traditional_kz": "4 (жақсы)", "traditional_ru": "4 (хорошо)"}
    if score >= 80:
        return {"letter": "B",  "gpa": 3.0,  "traditional_kz": "4 (жақсы)", "traditional_ru": "4 (хорошо)"}
    if score >= 75:
        return {"letter": "B-", "gpa": 2.67, "traditional_kz": "4 (жақсы)", "traditional_ru": "4 (хорошо)"}
    if score >= 70:
        return {"letter": "C+", "gpa": 2.33, "traditional_kz": "4 (жақсы)", "traditional_ru": "4 (хорошо)"}
    if score >= 65:
        return {"letter": "C",  "gpa": 2.0,  "traditional_kz": "3 (қанағат)", "traditional_ru": "3 (удовл)"}
    if score >= 60:
        return {"letter": "C-", "gpa": 1.67, "traditional_kz": "3 (қанағат)", "traditional_ru": "3 (удовл)"}
    if score >= 55:
        return {"letter": "D+", "gpa": 1.33, "traditional_kz": "3 (қанағат)", "traditional_ru": "3 (удовл)"}
    if score >= 50:
        return {"letter": "D",  "gpa": 1.0,  "traditional_kz": "3 (қанағат)", "traditional_ru": "3 (удовл)"}

    return {"letter": "F", "gpa": 0, "traditional_kz": "2 (қанағаттанарлықсыз)", "traditional_ru": "2 (неуд)"}


def parse_sheet(sheet):
    """Parse a single student sheet."""
    logging.info(f"Parsing sheet: {sheet.title}")
    students = []
    
    # 1. READ HEADER ROWS (1-6) for metadata
    # We need Row 5 (index 4) for Hours/Credits
    # We need Row 6 (index 5) for Year Info
    
    row_iter = sheet.iter_rows(min_row=1, values_only=True)
    rows_buffer = []
    
    # Read first 6 rows
    for _ in range(6):
        try:
            rows_buffer.append(next(row_iter))
        except StopIteration:
            break
            
    if len(rows_buffer) < 6:
        logging.warning("Sheet has fewer than 6 rows, skipping.")
        return []

    row_hours = rows_buffer[4] # Row 5
    row_years = rows_buffer[5] # Row 6

    # Extract Year Info
    year_enrollment = get_col_val(row_years, 150)
    year_graduation = get_col_val(row_years, 151)
    
    # BUILD SUBJECT METADATA (Hours/Credits) from Mappings + Row 5
    subject_metadata = {}
    all_mappings = MAPPING_PAGE1 + MAPPING_BASIC_MODULES + MAPPING_PROFESSIONAL + MAPPING_OTHER
    
    for subj_name, pct_col, _ in all_mappings:
        hours_raw = get_col_val(row_hours, pct_col)
        h, c = parse_hours_credits(str(hours_raw))
        subject_metadata[subj_name] = {"hours": h, "credits": c}
        
    # 2. PROCESS STUDENT ROWS (Row 7+)
    count = 0
    import random # Import locally if not at top

    for row in row_iter:
        # Check Name column (Col 2 / Index 1)
        name = get_col_val(row, 2)
        if not name:
            continue # Skip empty rows, but don't break immediately in case of gaps?
            # Typically verify if serial number (Col 1) exists
            if not get_col_val(row, 1):
                continue
        
        # Parse Dates
        start_y = str(year_enrollment).strip() if year_enrollment else "2023" # Default fallbacks
        end_y = str(year_graduation).strip() if year_graduation else "2026"
        
        # Format 2023.0 -> 2023
        if start_y.endswith('.0'): start_y = start_y[:-2]
        if end_y.endswith('.0'): end_y = end_y[:-2]
        
        # Normalize date if format is "01.09.2023" -> "2023"
        if len(start_y) > 4: start_y = start_y[-4:]
        if len(end_y) > 4: end_y = end_y[-4:]

        # Diploma ID: Use value -> Random 6-digit if empty (NO Prefix)
        dip_id = str(get_col_val(row, 152)) if get_col_val(row, 152) else ""
        if not dip_id or len(dip_id) < 3:
             dip_id = f"{random.randint(100000, 999999)}"

        student_data = {
            "full_name": name,
            "start_year": start_y,
            "end_year": end_y,
            "diploma_id": dip_id,
            "college_name": "Жамбыл инновациялық жоғары колледжінде",
            "specialization": "04110100 Есеп және аудит",
            "qualification": "4S04110102 Бухгалтер",
            "grades": {}
        }
        
        valid_grades_count = 0
        for subj_name, pct_col, _ in all_mappings:
            # ONLY READ PERCENTAGE (Points)
            pct = get_col_val(row, pct_col)
            
            # Calculate details based on Points
            details = calculate_grade_details(pct)
            
            # Metadata
            meta = subject_metadata.get(subj_name, {"hours": "", "credits": ""})
            
            # Check if this grade is valid (has points)
            if pct is not None:
                valid_grades_count += 1
                
            student_data["grades"][subj_name] = {
                "points": pct,
                "letter": details["letter"],
                "gpa": details["gpa"],
                "traditional": details["traditional_kz"], # Default to KZ for backward compat
                "traditional_ru": details["traditional_ru"], # specific for RU generator
                "hours": meta["hours"],
                "credits": meta["credits"]
            }
            
        students.append(student_data)
        count += 1
        
    logging.info(f"Extracted {count} students from {sheet.title}. (Last student had {valid_grades_count} valid grades)")
    return students

def parse_workbook(path):
    """Main entry: Parse all Accountant sheets."""
    if not os.path.exists(path):
        logging.error(f"File not found: {path}")
        return []
        
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_students = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("3D"):
            sheet = wb[sheet_name]
            students = parse_sheet(sheet)
            all_students.extend(students)
            
    return all_students

if __name__ == "__main__":
    # Test run
    s = parse_workbook(SOURCE_FILE)
    if s:
        print(f"Total parsed: {len(s)}")
        print("Sample Data (First Student):")
        import json
        # serialize with defaults
        def default_serializer(o):
            return str(o)
        print(json.dumps(s[0], indent=2, ensure_ascii=False, default=default_serializer))
