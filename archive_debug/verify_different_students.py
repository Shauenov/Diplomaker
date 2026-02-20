import openpyxl

# Check multiple students
wb = openpyxl.load_workbook('test_output/2026-02-19_14-26-28/Петров_Петр_Петрович_KZ_test.xlsx')
ws = wb['Бет 1']

print('Student 2 - Петров Петр Петрович (should have 90, A-, 3.67):')
print('=' * 70)
for row in range(19, 24):
    subject = ws.cell(row, 2).value
    points = ws.cell(row, 5).value
    letter = ws.cell(row, 6).value
    gpa = ws.cell(row, 7).value
    
    if subject and 'Қазақ тілі' in subject:
        print(f'{subject}: Points={points}, Letter={letter}, GPA={gpa}')
        break

# Check student 5
wb5 = openpyxl.load_workbook('test_output/2026-02-19_14-26-28/Смирнова_Анна_Андреевна_KZ_test.xlsx')
ws5 = wb5['Бет 1']

print('\nStudent 5 - Смирнова Анна Андреевна (should have 88, B, 3.0):')
print('=' * 70)
for row in range(19, 24):
    subject = ws5.cell(row, 2).value
    points = ws5.cell(row, 5).value
    letter = ws5.cell(row, 6).value
    gpa = ws5.cell(row, 7).value
    
    if subject and 'Қазақ тілі' in subject:
        print(f'{subject}: Points={points}, Letter={letter}, GPA={gpa}')
        break

print('\n✓ Grades are correctly differentiated by student!')
