# -*- coding: utf-8 -*-
"""
inspect_generated_diploma.py
Open a generated diploma and check which subjects have missing data
"""

import pandas as pd
import os
from openpyxl import load_workbook

# Get first generated diploma
output_dir = "Diplomas_Batch"
files = [f for f in os.listdir(output_dir) if f.endswith("_KZ.xlsx")]
if not files:
    print("No diplomas found!")
    exit(1)

diploma_path = os.path.join(output_dir, files[0])
print(f"Inspecting: {files[0]}\n")

# Read with openpyxl to preserve formatting and structure
wb = load_workbook(diploma_path)

# Typically page 1 and 2
for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"=== Sheet {sheet_idx}: {sheet_name} ===\n")
    
    # Columns: A=№, B=Subject, C=Hours, D=Credits, E=Points, F=Letter, G=GPA, H=Traditional
    # Read from row 1 to end
    
    empty_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not row or all(v is None for v in row[:8]):
            continue  # Skip completely empty rows
        
        idx, subject, hours, credits, points, letter, gpa, trad = row[:8]
        
        # Skip header rows and completely empty subjects
        if not subject or str(subject).strip() in ["Пән атауы", "Subject", ""]:
            continue
        
        subject_str = str(subject).strip()[:50] if subject else ""
        hours_str = str(hours) if hours not in [None, ""] else "[EMPTY]"
        credits_str = str(credits) if credits not in [None, ""] else "[EMPTY]"
        points_str = str(points) if points not in [None, ""] else "[EMPTY]"
        
        # Check if subject data is missing
        is_empty = hours_str == "[EMPTY]" and credits_str == "[EMPTY]" and points_str == "[EMPTY]"
        
        if is_empty:
            empty_count += 1
            print(f"❌ ROW {row_idx}: {subject_str}")
            print(f"    Hours={hours_str}, Credits={credits_str}, Points={points_str}\n")
    
    if empty_count == 0:
        print(f"✅ All subjects filled!\n")
    else:
        print(f"⚠ Found {empty_count} empty subjects\n")
