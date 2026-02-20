# -*- coding: utf-8 -*-
"""
fill_attestation_hours_v2.py
Add hours/credits for attestation row, handling merged cells
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SOURCE = "2025-2026 диплом бағалары (ТОЛЫҚ) қызыл диплом жазылған соңғысы точно (1).xlsx"

# Column and row indices
COL_ATTESTATION = 207  # Column 207 (need to check exact column)
ROW_HOURS = 4  # Excel row 4 (hours row)

HOURS_VALUE = "108с-4.5к"

sheets = ["3Ғ-1", "3Ғ-2", "3Ғ-3", "3Ғ-4"]

print(f"Filling attestation hours in source file:\n")

wb = load_workbook(SOURCE)

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        print(f"⚠ Sheet {sheet_name} not found!")
        continue
    
    ws = wb[sheet_name]
    
    # Convert column number to letter
    col_letter = get_column_letter(COL_ATTESTATION)
    cell_addr = f"{col_letter}{ROW_HOURS}"
    
    print(f"Processing {sheet_name}: {cell_addr}")
    
    # Check if cell is merged and unmerge if needed
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if cell_addr in merged_range:
            print(f"  >> Found merged range: {merged_range}")
            ws.unmerge_cells(str(merged_range))
            print(f"  >> Unmerged!")
    
    # Now write the value
    ws[cell_addr].value = HOURS_VALUE
    print(f"✅ Set {cell_addr} = '{HOURS_VALUE}'\n")

# Save
wb.save(SOURCE)
print(f"✅ Saved: {SOURCE}")
