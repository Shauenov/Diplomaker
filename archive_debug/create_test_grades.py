# -*- coding: utf-8 -*-
"""
create_test_grades.py
=====================
Creates a test copy of the source diploma grades file with realistic
filled-in percentage scores for all students, so the full pipeline
can be tested end-to-end.

Imports paths and thresholds from config.settings to stay DRY.

Run:
    python create_test_grades.py                    # default source & output
    python create_test_grades.py --seed 123         # custom seed
    python create_test_grades.py -o my_test.xlsx    # custom output path

Output: test_grades_filled.xlsx (same structure as the real source)
"""

import argparse
import random
import shutil
import openpyxl

from config.settings import SOURCE_FILE, TEST_GRADES_FILE, GRADE_THRESHOLDS
from config.programs import get_sheets_for_program

# Sheets to fill (from config — uses Cyrillic Ғ)
SHEETS = get_sheets_for_program("IT")

ROW_HEADER_P = 5      # Excel row 5 (1-based) = column labels (п/б/цэ/трад)
ROW_DATA_START = 6    # Excel row 6 (1-based) = first student

# Build realistic percentage pool from GRADE_THRESHOLDS
# Weight: A-range less frequent, B-range most frequent, C/D/F rare
_WEIGHTS = {
    95: 2,   # A
    90: 3,   # A-
    85: 4,   # B+
    80: 4,   # B
    75: 3,   # B-
    70: 2,   # C+
    65: 1,   # C
    60: 1,   # C-
    55: 1,   # D+
    50: 1,   # D
    0:  0,   # F — only injected explicitly below
}

SCORE_POOL: list[int] = []
thresholds_sorted = sorted(GRADE_THRESHOLDS.keys(), reverse=True)

for i, threshold in enumerate(thresholds_sorted):
    weight = _WEIGHTS.get(threshold, 1)
    if weight == 0:
        continue
    # Range: from this threshold to next one above (exclusive) or 100
    upper = 100 if i == 0 else thresholds_sorted[i - 1] - 1
    lower = threshold
    SCORE_POOL.extend(list(range(lower, upper + 1)) * weight)

# Add a few F scores (below 50) and explicit 0
SCORE_POOL.extend([0, 10, 25, 35, 42, 48])

# Fraction of grades to leave empty (simulates elective-not-taken)
EMPTY_FRACTION = 0.03


def _is_elective_column(ws, col: int) -> bool:
    """Check if a score column corresponds to an elective subject."""
    for check_row in (1, 2):
        val = ws.cell(row=check_row, column=col).value
        if val and any(tag in str(val).lower() for tag in ("факультатив", "ф1", "ф2", "ф3")):
            return True
    return False


def _score_to_letter(score: int) -> str:
    """Quick score-to-letter for reporting only."""
    for threshold in sorted(GRADE_THRESHOLDS.keys(), reverse=True):
        if score >= threshold:
            return GRADE_THRESHOLDS[threshold]["letter"]
    return "F"


def main():
    parser = argparse.ArgumentParser(description="Create test grades file")
    parser.add_argument(
        "-s", "--source",
        default=SOURCE_FILE,
        help="Path to source Excel file (default: config.settings.SOURCE_FILE)",
    )
    parser.add_argument(
        "-o", "--output",
        default=str(TEST_GRADES_FILE),
        help="Path to output test file (default: config.settings.TEST_GRADES_FILE)",
    )
    parser.add_argument(
        "--seed", type=int, default=42,
        help="Random seed for reproducibility (default: 42)",
    )
    args = parser.parse_args()

    random.seed(args.seed)

    print(f"Copying {args.source} -> {args.output} ...")
    shutil.copy2(args.source, args.output)

    wb = openpyxl.load_workbook(args.output)  # data_only=False to preserve formulas

    total_students = 0
    total_grades = 0
    total_empty = 0
    grade_distribution: dict[str, int] = {}

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [SKIP] Sheet '{sheet_name}' not found.")
            continue

        ws = wb[sheet_name]
        print(f"  Filling sheet: {sheet_name}")

        # Find all 'п' (points) columns — header row 5 has 'п' in those cells
        point_cols = []
        header_row = ws[ROW_HEADER_P]
        for cell in header_row:
            if cell.value == "п":
                point_cols.append(cell.column)  # 1-based

        if not point_cols:
            print(f"    [WARN] No 'п' columns found in header row {ROW_HEADER_P}")
            continue

        print(f"    Found {len(point_cols)} score columns.")

        # Fill each student row
        students_filled = 0
        for row_idx in range(ROW_DATA_START, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=2)
            if not name_cell.value or str(name_cell.value).strip() == "":
                continue

            students_filled += 1
            for col in point_cols:
                # Randomly leave some elective grades empty
                if random.random() < EMPTY_FRACTION and _is_elective_column(ws, col):
                    ws.cell(row=row_idx, column=col).value = None
                    total_empty += 1
                    continue

                score = random.choice(SCORE_POOL)
                ws.cell(row=row_idx, column=col).value = score
                total_grades += 1

                # Track distribution
                letter = _score_to_letter(score)
                grade_distribution[letter] = grade_distribution.get(letter, 0) + 1

        total_students += students_filled
        print(f"    Filled {students_filled} students.")

    wb.save(args.output)

    # Summary report
    print(f"\nSaved: {args.output}")
    print(f"  Total students:  {total_students}")
    print(f"  Total grades:    {total_grades}")
    print(f"  Empty (elective): {total_empty}")
    if total_grades:
        print(f"\nGrade distribution:")
        for letter in ["A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "F"]:
            count = grade_distribution.get(letter, 0)
            pct = (count / total_grades * 100) if total_grades else 0
            bar = "#" * int(pct / 2)
            print(f"    {letter:>3s}: {count:5d} ({pct:5.1f}%) {bar}")

    print(f"\nRun:  python -m batch._generate_it")


if __name__ == "__main__":
    main()
