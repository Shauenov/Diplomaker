"""
Comprehensive verification script:
1. Reads subjects from Excel source (all sheets)
2. Reads subjects from generated diplomas (KZ+RU)
3. Compares and reports mismatches
"""
import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import re
import json

SOURCE_FILE = r"c:\Users\user\OneDrive\Рабочий стол\template\2025-2026 диплом бағалары (ТОЛЫҚ) қызыл диплом жазылған соңғысы точно (1).xlsx"
DIPLOMAS_DIR = r"c:\Users\user\OneDrive\Рабочий стол\template\Diplomas_Batch"

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


def normalize(text):
    """Normalize for comparison: lowercase, strip spaces/dots."""
    if not text:
        return ''
    t = text.strip().lower()
    t = t.replace('.', '').replace(',', '').replace(':', '')
    t = re.sub(r'\s+', '', t)
    t = re.sub(r'([а-яёa-z]+)0*([1-9]+)', r'\1\2', t)
    return t


# ─ Step 1: Read Excel source ─────────────────────────────────────────────────
print("=" * 70)
print("STEP 1: Reading source Excel...")
print("=" * 70)

wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
sheets = wb.sheetnames
print(f"Sheets: {sheets}")

excel_subjects = {}  # sheet_name -> set of normalized subject names

for sheet_name in sheets:
    ws = wb[sheet_name]
    subjects = set()
    
    # Read all non-empty cells from column B (subject names column)
    # Subject rows typically in col 2 (B)
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val and isinstance(val, str) and len(val.strip()) > 3:
            stripped = val.strip()
            # Skip obvious header values
            if any(skip in stripped.lower() for skip in ['атауы', 'название', 'пәні', '#', '№']):
                continue
            subjects.add(stripped)
    
    excel_subjects[sheet_name] = subjects
    print(f"  {sheet_name}: {len(subjects)} subjects found")

wb.close()


# ─ Step 2: Read diploma subjects ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 2: Reading diploma subjects...")
print("=" * 70)

def get_shared_strings(fpath):
    """Extract shared strings from xlsx."""
    with zipfile.ZipFile(fpath) as z:
        if 'xl/sharedStrings.xml' not in z.namelist():
            return []
        with z.open('xl/sharedStrings.xml') as f:
            root = ET.parse(f).getroot()
            return [''.join(t.text or '' for t in si.iter(f'{{{NS}}}t'))
                    for si in root.findall(f'{{{NS}}}si')]


diploma_subjects = {}  # group -> {lang -> list of subjects}

for fname in sorted(os.listdir(DIPLOMAS_DIR)):
    if not fname.endswith('.xlsx'):
        continue
    
    # Parse filename: "3F-1_Студент Имя_KZ.xlsx"
    parts = fname.replace('.xlsx', '').split('_')
    if len(parts) < 2:
        continue
    group = parts[0]  # e.g. "3F-1" or "3D-1"
    lang = parts[-1]  # "KZ" or "RU"
    
    if group not in diploma_subjects:
        diploma_subjects[group] = {'KZ': set(), 'RU': set()}
    
    # Only process one file per group per lang (they all have same subjects)
    if diploma_subjects[group][lang]:
        continue
    
    try:
        fpath = os.path.join(DIPLOMAS_DIR, fname)
        strings = get_shared_strings(fpath)
        
        subj_set = set()
        for s in strings:
            s = s.strip()
            # Filter: keep lines that look like subject names
            # Must be >3 chars, not a number, not a year, not a grade
            if (len(s) > 3 
                and not re.match(r'^\d+([.,]\d+)?$', s)  # not pure number
                and not re.match(r'^\d{4}$', s)  # not year
                and s not in ['A', 'B', 'B+', 'C+', 'C', 'D+', 'D', 'F']
                and not s.startswith('зачтено')
                and not s.startswith('сынақ')
                and not s.startswith('Отлично')
                and not s.startswith('Хорошо')
                ):
                subj_set.add(s)
        
        diploma_subjects[group][lang] = subj_set
        
    except Exception as e:
        print(f"  ERROR reading {fname}: {e}")

print(f"Groups found in diplomas: {list(diploma_subjects.keys())}")


# ─ Step 3: Compare ───────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 3: Comparing subjects...")
print("=" * 70)

# Map diploma group to excel sheet
GROUP_TO_SHEET = {
    '3D-1': '3D-1',
    '3D-2': '3D-2',
    '3F-1': '3Ғ-1',
    '3F-2': '3Ғ-2',
    '3F-3': '3Ғ-3',
    '3F-4': '3Ғ-4',
}

report = {}

for group, langs in sorted(diploma_subjects.items()):
    sheet = GROUP_TO_SHEET.get(group)
    if not sheet or sheet not in excel_subjects:
        print(f"  Skipping {group} (no matching sheet)")
        continue
    
    excel_set = excel_subjects[sheet]
    excel_norm = {normalize(s): s for s in excel_set}
    
    report[group] = {}
    
    for lang, dipl_subjects in langs.items():
        if not dipl_subjects:
            continue
        
        dipl_norm = {normalize(s): s for s in dipl_subjects}
        
        # Find subjects in diploma but NOT in excel
        in_dipl_not_excel = []
        for nk, orig in sorted(dipl_norm.items()):
            if nk not in excel_norm:
                in_dipl_not_excel.append(orig)
        
        # Find subjects in excel but NOT in diploma
        in_excel_not_dipl = []
        for nk, orig in sorted(excel_norm.items()):
            if nk not in dipl_norm:
                in_excel_not_dipl.append(orig)
        
        print(f"\n  [{group} / {lang}]")
        if not in_dipl_not_excel and not in_excel_not_dipl:
            print(f"    ✅ PERFECT MATCH ({len(dipl_norm)} subjects)")
        else:
            if in_dipl_not_excel:
                print(f"    ⚠️  In diploma but NOT in Excel ({len(in_dipl_not_excel)}):")
                for s in in_dipl_not_excel[:10]:
                    print(f"      - {s[:80]}")
            if in_excel_not_dipl:
                print(f"    ⚠️  In Excel but NOT in diploma ({len(in_excel_not_dipl)}):")
                for s in in_excel_not_dipl[:10]:
                    print(f"      - {s[:80]}")
        
        report[group][lang] = {
            'in_diploma_not_excel': in_dipl_not_excel,
            'in_excel_not_diploma': in_excel_not_dipl,
            'ok': not in_dipl_not_excel and not in_excel_not_dipl,
        }

# ─ Summary ───────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("SUMMARY")
print("=" * 70)
total_ok = 0
total_issues = 0
for group, langs in sorted(report.items()):
    for lang, res in langs.items():
        if res.get('ok'):
            print(f"  ✅ {group} {lang}")
            total_ok += 1
        else:
            n1 = len(res.get('in_diploma_not_excel', []))
            n2 = len(res.get('in_excel_not_diploma', []))
            print(f"  ❌ {group} {lang} — {n1} extra in diploma, {n2} missing from diploma")
            total_issues += 1

print(f"\nTotal OK: {total_ok}, Issues: {total_issues}")

# Save full report
with open('verification_report.json', 'w', encoding='utf-8') as f:
    json.dump(report, f, ensure_ascii=False, indent=2)
print("\nFull report saved to verification_report.json")
