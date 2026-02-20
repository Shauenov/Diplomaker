
import openpyxl
import random
import os

SOURCE = r'c:\Users\user\OneDrive\Рабочий стол\template\2025-2026 диплом бағалары (ТОЛЫҚ) қызыл диплом жазылған соңғысы точно (1).xlsx'
OUTPUT = r'c:\Users\user\OneDrive\Рабочий стол\template\2025-2026_FILLED.xlsx'

# Mapping taken from parse_grades.py
# (Subject Name, Pct Col Index) - 1-based index
# Note: In openpyxl, cell(row, col) is 1-based.
# parse_grades uses get_col_val(row, idx) where idx likely matches the Excel column index.

MAPPINGS = [
    # Page 1
    (8, 7), (11, 7), (14, 7), (17, 7), (20, 7), (23, 7), (26, 7), (29, 7), 
    (32, 7), (35, 7), (38, 7), (41, 7), (44, 7),
    # Base Modules
    (47, 7), (50, 7), (53, 7), (56, 7),
    # Prof Modules
    (59, 7), # PM 1 Header (skip?)
    (74, 7), (77, 7), (80, 7), (83, 7),
    (86, 7), # PM 2 Header
    (88, 7), (91, 7), (94, 7), (97, 7),
    (100, 7), # PM 3 Header
    (102, 7), (105, 7), (108, 7), (111, 7),
    (114, 7), # PM 4 Header
    (116, 7), (119, 7), (122, 7), (125, 7), (128, 7), (131, 7), (134, 7),
    (137, 7), # PM 5 Header
    (139, 7), (142, 7),
    # Other
    (145, 7), (148, 7)
]

# We need to be careful. The MAPPINGS above are (Col Index, Row Index of Metadata) or something?
# Let's check parse_grades.py mapping structure.
# MAPPING_PAGE1 = [ ("Казахский язык", 8, None), ... ]
# The integer is the COLUMN INDEX for percentage.

# Let's redefine based on parse_grades.py constants
COLS_TO_FILL = [
    8, 11, 14, 17, 20, 23, 26, 29, 32, 35, 38, 41, 44,
    47, 50, 53, 56,
    # PM1
    # 59 is PM1 header (points null)
    # RO 1.1...
    # The parser iterates: for subj_name, pct_col, trad_col in all_mappings:
    # We need to fill pct_col (N), Letter (N+1), GPA (N+2), Trad (N+3)
    
    # Let's just blindly fill columns that look like grade slots in rows 7+
    # We can infer from the header in row 6: "п" (percent), "ә" (letter), "б" (gpa), "т" (trad)
]

def derive_grade(score):
    if score >= 95: return "A", 4.0, "өте жақсы"
    if score >= 90: return "A-", 3.67, "өте жақсы"
    if score >= 85: return "B+", 3.33, "жақсы"
    if score >= 80: return "B", 3.0, "жақсы"
    if score >= 75: return "B-", 2.67, "жақсы"
    if score >= 70: return "C+", 2.33, "қанағат"
    if score >= 65: return "C", 2.0, "қанағат"
    if score >= 60: return "C-", 1.67, "қанағат"
    if score >= 55: return "D+", 1.33, "қанағат"
    if score >= 50: return "D", 1.0, "қанағат"
    return "F", 0, "қанағатсыз"

def fill_workbook():
    print(f"Loading {SOURCE}...")
    wb = openpyxl.load_workbook(SOURCE)
    
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("3D"):
            continue
            
        print(f"Processing {sheet_name}...")
        ws = wb[sheet_name]
        
        # Row 6 has headers. Let's find columns that have "п" (Percentage) in row 6
        grade_cols = []
        for col in range(5, 200): # Scan columns
            val = ws.cell(row=6, column=col).value
            if val and isinstance(val, str) and "п" in val.lower().strip():
                grade_cols.append(col)
        
        print(f"Found {len(grade_cols)} grade columns.")
        
        # Fill rows 7 to 40 (student rows)
        # Using max_row is risky if sheet has trailing empty rows
        count = 0
        for row in range(7, 60): 
            # Check if name exists in col 2
            name = ws.cell(row=row, column=2).value
            if not name:
                continue
                
            count += 1
            # Fill grades
            for col in grade_cols:
                # Generate random score 70-100
                score = random.randint(70, 99)
                letter, gpa, trad = derive_grade(score)
                
                # Write Percentage
                ws.cell(row=row, column=col).value = score
                # Write Letter (next col)
                ws.cell(row=row, column=col+1).value = letter
                # Write GPA (next next)
                ws.cell(row=row, column=col+2).value = gpa
                # Write Trad (next next next) - roughly
                # Check if col+3 header is 'трад'
                # ws.cell(row=row, column=col+3).value = trad 
                # (Actually parser derives trad from pct if missing, but let's fill it to be safe or leave blank to test derivation)
                # Let's derive it in parser.
                
    print(f"Saving to {OUTPUT}...")
    wb.save(OUTPUT)
    print("Done.")

if __name__ == "__main__":
    fill_workbook()
