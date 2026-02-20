import openpyxl
import os

OUTPUT_DIR = "Diplomas_Batch"

def verify():
    print(f"Scanning {OUTPUT_DIR}...")
    files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx") and not f.startswith("~$")]
    
    kz_count = 0
    ru_count = 0
    kz_ok = 0
    ru_ok = 0
    
    for f in files:
        path = os.path.join(OUTPUT_DIR, f)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            # Check Page 4
            if "Бет 4" in wb.sheetnames:
                ws = wb["Бет 4"]
            elif "Page 4" in wb.sheetnames:
                ws = wb["Page 4"] # RU might differ?
            else:
                # RU generator logic? RU uses "Стр 1", "Стр 2"...?
                # Let's check sheet names for RU
                if "_RU" in f:
                    # In generate_diploma_it_ru.py, sheet names are likely "Стр 1" etc.
                    # Or check for any sheet
                    ws = wb[wb.sheetnames[-1]] # Last sheet?
                else:
                    ws = wb[wb.sheetnames[-1]]

            found_grade = False
            is_kz = "_KZ" in f
            is_ru = "_RU" in f
            target_grade = "сынақ" if is_kz else "зачтено"
            
            # Scan first 20 rows of the sheet (page 4 subjects)
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
                vals = [c.value for c in row]
                line = " | ".join([str(x) if x else "" for x in vals])
                if "Факультатив" in line:
                    # Check if target_grade is in line
                    # Note: "сынақ" might be "сынак" or similar if encoding issues, but we expect exact match
                    if target_grade in line:
                        found_grade = True
            
            if is_kz:
                kz_count += 1
                if found_grade: kz_ok += 1
                else: print(f"FAIL KZ: {f} (Missing '{target_grade}')")
            elif is_ru:
                ru_count += 1
                if found_grade: ru_ok += 1
                else: print(f"FAIL RU: {f} (Missing '{target_grade}')")
                
            wb.close()
            
        except Exception as e:
            print(f"Error reading {f}: {e}")

    print(f"KZ: {kz_ok}/{kz_count} ok")
    print(f"RU: {ru_ok}/{ru_count} ok")

if __name__ == "__main__":
    verify()
