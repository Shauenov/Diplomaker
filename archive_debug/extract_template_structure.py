import openpyxl
import os

def extract_subjects(file_path):
    print(f"--- Extracting from {os.path.basename(file_path)} ---")
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name}")
        subjects = []
        # Assuming subjects are in Column B (index 2), starting from some row
        # We'll scan column B
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if not row or len(row) < 2:
                continue
            
            val_a = row[0] # Index
            val_b = row[1] # Subject
            
            # Simple heuristic: if A is a number or B looks like a subject
            if val_b and str(val_b).strip():
                # Filter out obvious headers if possible, or just print everything reasonable
                # We'll print everything that looks like content to be safe
                if isinstance(val_a, int) or (isinstance(val_a, str) and val_a.isdigit()):
                     print(f"  Row {row[0]}: {val_b}")
                elif "Пән атауы" in str(val_b) or "Subject" in str(val_b) or "Наименование" in str(val_b):
                     print(f"  [HEADER] {val_b}")
                else:
                     # Some subjects might not have an index yet?
                     # Just print likely subjects
                     if len(str(val_b)) > 5: 
                        print(f"  Row {row[0] if row[0] else '?'} : {val_b}")

if __name__ == "__main__":
    base_path = r"c:\Users\user\OneDrive\Рабочий стол\template"
    templates = [
        "Diplom_IT_KZ_Template.xlsx",
        "Diplom_IT_RU_Template.xlsx"
    ]
    
    for t in templates:
        path = os.path.join(base_path, t)
        if os.path.exists(path):
            extract_subjects(path)
        else:
            print(f"File not found: {path}")
