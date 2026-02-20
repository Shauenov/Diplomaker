# -*- coding: utf-8 -*-
"""
fill_attestation_hours.py
Add hours/credits for attestation row in source Excel
"""

import pandas as pd
from openpyxl import load_workbook

SOURCE = "2025-2026 диплом бағалары (ТОЛЫҚ) қызыл диплом жазылған соңғысы точно (1).xlsx"

# Column and row indices (0-based for pandas)
COL_ATTESTATION = 206  # Column 206 for attestation
ROW_HOURS = 3  # Row 4 in Excel = index 3 in pandas

HOURS_VALUE = "108с-4.5к"  # 108 hours, 4.5 credits

sheets = ["3Ғ-1", "3Ғ-2", "3Ғ-3", "3Ғ-4"]

print(f"Filling attestation hours in source file:\n")

# Use openpyxl to preserve formatting
wb = load_workbook(SOURCE)

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        print(f"⚠ Sheet {sheet_name} not found!")
        continue
    
    ws = wb[sheet_name]
    
    # Convert 0-based col index to Excel column letter
    # Col 206 = GL (since A=1, Z=26, AA=27, ..., AZ=52, BA=53, ..., GL=206)
    # Calculate: 206 = 7*26 + 24, so it's GL
    col_letter = ""
    col_num = COL_ATTESTATION
    while col_num > 0:
        col_num -= 1
        col_letter = chr(col_num % 26 + 65) + col_letter
        col_num //= 26
    
    cell_addr = f"{col_letter}{ROW_HOURS + 1}"  # Excel uses 1-based row numbering
    
    ws[cell_addr].value = HOURS_VALUE
    print(f"✅ {sheet_name}: Set {cell_addr} = '{HOURS_VALUE}'")

# Save the workbook
wb.save(SOURCE)
print(f"\n✅ Source file updated: {SOURCE}")
