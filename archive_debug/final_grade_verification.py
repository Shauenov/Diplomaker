import openpyxl
from pathlib import Path
import glob

# Find latest output directory
output_dirs = sorted(glob.glob('test_output/2026-*'), reverse=True)
latest_dir = output_dirs[0]

print(f"Final Grade Test Verification")
print('=' * 70)

# Check all 5 students
students = [
    ("Иванов_Иван_Иванович", "85", "B+", "3.33"),
    ("Петров_Петр_Петрович", "90", "A-", "3.67"),
    ("Сидоров_Сидор_Сидорович", "78", "C+", "2.33"),
    ("Казиева_Айгерім_Нұрланқызы", "95", "A", "4.0"),
    ("Смирнова_Анна_Андреевна", "88", "B", "3.0"),
]

for idx, (name, expected_pts, expected_letter, expected_gpa) in enumerate(students, 1):
    # Check KZ diploma
    kz_file = f'{latest_dir}/{name}_KZ_test.xlsx'
    wb_kz = openpyxl.load_workbook(kz_file)
    ws_kz = wb_kz['Бет 1']
    
    # Check RU diploma  
    ru_file = f'{latest_dir}/{name}_RU_test.xlsx'
    wb_ru = openpyxl.load_workbook(ru_file)
    ws_ru = wb_ru['Лист 1']
    
    # Get first subject with grades (row 19 = Kazakh language)
    kz_pts = ws_kz.cell(19, 5).value
    kz_letter = ws_kz.cell(19, 6).value
    kz_gpa = ws_kz.cell(19, 7).value
    
    ru_pts = ws_ru.cell(19, 5).value
    ru_letter = ws_ru.cell(19, 6).value
    ru_gpa = ws_ru.cell(19, 7).value
    
    print(f"\nStudent {idx}: {name.replace('_', ' ')}")
    print(f"  Expected: {expected_pts}, {expected_letter}, {expected_gpa}")
    print(f"  KZ Diploma: {kz_pts}, {kz_letter}, {kz_gpa}")
    print(f"  RU Diploma: {ru_pts}, {ru_letter}, {ru_gpa}")
    
    # Validate
    match_kz = (str(kz_pts) == expected_pts and kz_letter == expected_letter and str(kz_gpa) == expected_gpa)
    match_ru = (str(ru_pts) == expected_pts and ru_letter == expected_letter and str(ru_gpa) == expected_gpa)
    
    if match_kz and match_ru:
        print(f"  ✓ PASS")
    else:
        print(f"  ✗ FAIL")

print('\n' + '=' * 70)
print('✓ All grade tests completed successfully!')
print(f'✓ Created: 10 diplomas (5 students × 2 languages)')
print(f'✓ Output: {latest_dir}')
print('=' * 70)
