"""Quick check: what sheets exist in the grades file and what data is in them."""
import openpyxl

SOURCE_FILE = r"c:\Users\user\OneDrive\Рабочий стол\template\2025-2026 диплом бағалары (ТОЛЫҚ) қызыл диплом жазылған соңғысы точно (1).xlsx"

print("Opening file...")
wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

# Check F-group sheets
for sheet_name in wb.sheetnames:
    if 'Ғ' in sheet_name or 'F' in sheet_name or 'ф' in sheet_name.lower():
        print(f"\n=== Sheet: {sheet_name} ===")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        for i, row in enumerate(rows):
            # Show first 20 values
            vals = [str(v)[:20] if v is not None else '-' for v in row[:20]]
            print(f"  Row {i+1}: {vals}")

wb.close()
print("\nDone.")
