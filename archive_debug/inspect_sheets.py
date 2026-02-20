import openpyxl
import os

files = [
    "Diplom_IT_KZ_Template.xlsx",
    "Diplom_IT_RU_Template.xlsx",
    "Diplomas_Batch/Аймахан Балауса Абайханқызы_KZ.xlsx",
    "Diplomas_Batch/Аймахан Балауса Абайханқызы_RU.xlsx"
]

base_path = r"c:\Users\user\OneDrive\Рабочий стол\template"

for f in files:
    path = os.path.join(base_path, f)
    print(f"\n--- {f} ---")
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"  {sheet}: max_row={ws.max_row}")
    except Exception as e:
        print(f"  Error: {e}")
