# -*- coding: utf-8 -*-
"""
verify_attestation_electives.py
Verify that attestation and electives have correct hours/credits
"""

import os
from openpyxl import load_workbook

diploma_dir = "Diplomas_Batch"
files = [f for f in os.listdir(diploma_dir) if f.endswith("_KZ.xlsx")]
if not files:
    print("No diplomas found!")
    exit(1)

diploma_path = os.path.join(diploma_dir, files[0])
print(f"Verifying: {files[0]}\n")

wb = load_workbook(diploma_path)
ws = wb["Бет 4"]  # Page 4 where attestation and electives are

print("=== Page 4 Details ===\n")

# Read columns B (subject), C (hours), D (credits)
for row_idx in range(1, ws.max_row + 1):
    subject_cell = ws.cell(row_idx, 2)  # Column B
    hours_cell = ws.cell(row_idx, 3)    # Column C
    credits_cell = ws.cell(row_idx, 4)  # Column D
    
    subject = subject_cell.value
    hours = hours_cell.value
    credits = credits_cell.value
    
    if not subject:
        continue
    
    subject_str = str(subject).strip()[:50]
    
    # Focus on attestation and electives
    if ("аттест" in subject_str.lower() or 
        "Факультатив" in subject_str or
        "факультативtik" in subject_str.lower() or
        "Ф1" in subject_str or "Ф2" in subject_str or "Ф3" in subject_str):
        
        print(f"Subject: {subject_str}")
        print(f"  Hours: {hours}, Credits: {credits}")
        
        if hours and credits:
            print(f"  ✅ FILLED\n")
        else:
            print(f"  ❌ EMPTY\n")
