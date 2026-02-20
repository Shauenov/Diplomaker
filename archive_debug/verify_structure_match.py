import openpyxl
import os

TEMPLATES = {
    "KZ": "Diplom_IT_KZ_Template.xlsx",
    "RU": "Diplom_IT_RU_Template.xlsx"
}

GENERATED_DIR = "Diplomas_Batch"

def get_subjects_from_sheet(ws):
    subjects = []
    # Iterate Column B (index 2)
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if not row: continue
        val = row[1]
        if val and str(val).strip():
            subjects.append(str(val).strip())
    return subjects

def verify_file(generated_path, template_path):
    print(f"Comparing {os.path.basename(generated_path)} \n      vs {os.path.basename(template_path)}")
    
    try:
        wb_gen = openpyxl.load_workbook(generated_path, read_only=True, data_only=True)
        wb_tmpl = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading files: {e}")
        return False

    sheets_gen = wb_gen.sheetnames
    sheets_tmpl = wb_tmpl.sheetnames
    
    if len(sheets_gen) != len(sheets_tmpl):
        print(f" FAIL: Sheet count mismatch ({len(sheets_gen)} vs {len(sheets_tmpl)})")
        return False

    all_good = True
    for i, sheet_name in enumerate(sheets_tmpl):
        ws_gen = wb_gen[sheets_gen[i]] # Use index matching in case names differ slightly
        ws_tmpl = wb_tmpl[sheet_name]
        
        subj_gen = get_subjects_from_sheet(ws_gen)
        subj_tmpl = get_subjects_from_sheet(ws_tmpl)
        
        if len(subj_gen) != len(subj_tmpl):
            print(f" FAIL: {sheet_name} subject count mismatch ({len(subj_gen)} vs {len(subj_tmpl)})")
            # print differences
            import difflib
            # diff = difflib.unified_diff(subj_gen, subj_tmpl, fromfile='Generated', tofile='Template', lineterm='')
            # for line in diff:
            #     print(line)
            
            print("   Generated:")
            for s in subj_gen: print(f"     - {s}")
            print("   Template:")
            for s in subj_tmpl: print(f"     + {s}")
            
            all_good = False
        else:
            # check content
            mismatch = False
            for j in range(len(subj_gen)):
                if subj_gen[j] != subj_tmpl[j]:
                    # Ignore Header placeholders
                    if "{diploma" in subj_tmpl[j] or "{full" in subj_tmpl[j]:
                        continue
                        
                    print(f" FAIL: {sheet_name} Row {j}: '{subj_gen[j]}' != '{subj_tmpl[j]}'")
                    mismatch = True
                    break
            if not mismatch:
                print(f" OK: {sheet_name}")
            else:
                all_good = False
                
    wb_gen.close()
    wb_tmpl.close()
    return all_good

def main():
    base_path = r"c:\Users\user\OneDrive\Рабочий стол\template"
    
    # Pick one KZ and one RU file to test
    kz_file = None
    ru_file = None
    
    gen_dir = os.path.join(base_path, GENERATED_DIR)
    for f in os.listdir(gen_dir):
        if f.startswith("~$"): continue
        if f.endswith("_KZ.xlsx") and not kz_file:
            kz_file = os.path.join(gen_dir, f)
            print(f"DEBUG: Selected KZ file: {kz_file}")
        if f.endswith("_RU.xlsx") and not ru_file:
            ru_file = os.path.join(gen_dir, f)
            print(f"DEBUG: Selected RU file: {ru_file}")
            
    if kz_file:
        template_kz = os.path.join(base_path, TEMPLATES["KZ"])
        verify_file(kz_file, template_kz)
        
    if ru_file:
        template_ru = os.path.join(base_path, TEMPLATES["RU"])
        verify_file(ru_file, template_ru)

if __name__ == "__main__":
    main()
