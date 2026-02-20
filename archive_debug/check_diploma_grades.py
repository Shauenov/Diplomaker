import openpyxl
from pathlib import Path
import glob

# Find latest output directory
output_dirs = sorted(glob.glob('test_output/2026-*'), reverse=True)
if not output_dirs:
    print("No test output directories found!")
    exit(1)

latest_dir = output_dirs[0]
print(f"Checking: {latest_dir}")
print('=' * 70)

# Check KZ diploma
kz_file = f'{latest_dir}/Иванов_Иван_Иванович_KZ_test.xlsx'
wb_kz = openpyxl.load_workbook(kz_file)
ws_kz = wb_kz['Бет 1']

print('\nKZ Diploma (Бет 1) - First 10 subjects:')
for row in range(19, 29):
    subject = ws_kz.cell(row, 2).value
    points = ws_kz.cell(row, 5).value
    letter = ws_kz.cell(row, 6).value
    
    if subject:
        pts_str = str(points) if points else "-"
        letter_str = str(letter) if letter else "-"
        print(f'  {subject:<35} | {pts_str:<6} | {letter_str}')

# Check RU diploma
ru_file = f'{latest_dir}/Иванов_Иван_Иванович_RU_test.xlsx'
wb_ru = openpyxl.load_workbook(ru_file)
ws_ru = wb_ru['Лист 1']

print('\nRU Diploma (Лист 1) - First 10 subjects:')
for row in range(19, 29):
    subject = ws_ru.cell(row, 2).value
    points = ws_ru.cell(row, 5).value
    letter = ws_ru.cell(row, 6).value
    
    if subject:
        pts_str = str(points) if points else "-"
        letter_str = str(letter) if letter else "-"
        print(f'  {subject:<35} | {pts_str:<6} | {letter_str}')

print('\n' + '=' * 70)
print('✓ Grade tests complete!')

