import openpyxl

SOURCE_FILE = "2025-2026 диплом бағалары (ТОЛЫҚ) қызыл диплом жазылған соңғысы точно (1).xlsx"
OUTPUT_FILE = "mock_filled_grades.xlsx"
SHEET_NAME = "3Ғ-1"

def fill_grades():
    print(f"Loading {SOURCE_FILE}...")
    wb = openpyxl.load_workbook(SOURCE_FILE)
    ws = wb[SHEET_NAME]
    
    max_col = ws.max_column
    
    # Fill Rows 10 to 32
    for row_idx in range(10, 33):
        name = ws.cell(row=row_idx, column=2).value
        # Even if name is empty, we might want to fill for testing, 
        # but let's stick to student rows.
        
        for col_idx in range(3, max_col + 1):
            offset = (col_idx - 3) % 4
            if offset == 0: ws.cell(row=row_idx, column=col_idx, value=95)
            elif offset == 1: ws.cell(row=row_idx, column=col_idx, value="A")
            elif offset == 2: ws.cell(row=row_idx, column=col_idx, value=4.0)
            elif offset == 3: ws.cell(row=row_idx, column=col_idx, value="өте жақсы")
            
    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done.")

if __name__ == "__main__":
    fill_grades()
