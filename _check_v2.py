"""Quick check: verify no numbers in column A for rows 1-14, and subjects start at row 15 with numbering from 1."""
import openpyxl
from pathlib import Path

for f in Path("test_output_v2").glob("*.xlsx"):
    wb = openpyxl.load_workbook(f)
    ws = wb.worksheets[0]
    print(f"\n=== {f.name} ===")
    
    # Check rows 1-14: column A should be empty
    a_empty = True
    for r in range(1, 15):
        v = ws.cell(r, 1).value
        if v is not None:
            print(f"  ROW {r} A={v} (should be empty!)")
            a_empty = False
    if a_empty:
        print("  Rows 1-14 col A: EMPTY (correct)")
    
    # Check first subject row
    for r in range(15, 20):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if b:
            print(f"  Row {r}: A={a}, B={str(b)[:40]}")
    
    wb.close()
