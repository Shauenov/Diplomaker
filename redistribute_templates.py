#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Redistribute subjects across template pages for even visual balance.

IT Templates (F): 13+21+21+10 → 17+17+17+14
  - Move 4 БМ from page 2 to page 1
  - Move КМ 09 + ОН 9.1-9.3 from page 3 to page 4

D Templates: already balanced (17+12+12+9), no changes needed.
"""

import copy
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


TEMPLATES_DIR = Path("templates")

# IT templates to modify
IT_TEMPLATES = [
    TEMPLATES_DIR / "Diplom_F_KZ_Template (4).xlsx",
    TEMPLATES_DIR / "Diplom_F_RU_Template (4).xlsx",
]

# New distribution for IT: which subjects go on which page
# Page 1: rows 15..31 (17 subjects: 13 general + 4 БМ)
# Page 2: rows 2..18 (17 subjects: КМ01-04 with ОН)
# Page 3: rows 2..18 (17 subjects: КМ05-08 with ОН)
# Page 4: rows 2..15 (14 subjects: КМ09-10 + practice + attest + 3 electives)
IT_NEW_DIST = [17, 17, 17, 14]  # subjects per page


def backup_file(filepath: Path):
    """Create backup before modifying."""
    backup = filepath.with_suffix('.xlsx.bak')
    shutil.copy2(filepath, backup)
    print(f"  Backup: {backup.name}")


def collect_subjects(wb) -> list:
    """Collect all subjects from all sheets into a flat list.
    
    Returns list of dicts: {value: str, row_height: float}
    """
    subjects = []
    for ws in wb.worksheets:
        sheet_name = ws.title
        # Page 1: subjects start at row 15 (rows 1-14 = student header)
        # Pages 2-4: subjects start at row 2
        start_row = 15 if ws == wb.worksheets[0] else 2
        
        for row in range(start_row, ws.max_row + 1):
            cell_b = ws.cell(row, 2)
            if cell_b.value and str(cell_b.value).strip():
                subjects.append({
                    'value': cell_b.value,
                    'row_height': ws.row_dimensions[row].height,
                })
    return subjects


def get_reference_style(ws, ref_row: int) -> dict:
    """Extract cell styles from a reference subject row."""
    styles = {}
    for col in range(1, 9):  # A through H
        cell = ws.cell(ref_row, col)
        styles[col] = {
            'font': copy.copy(cell.font),
            'alignment': copy.copy(cell.alignment),
            'border': copy.copy(cell.border),
            'fill': copy.copy(cell.fill),
            'number_format': cell.number_format,
        }
    return styles


def apply_style(ws, row: int, styles: dict):
    """Apply saved styles to a row."""
    for col, style in styles.items():
        cell = ws.cell(row, col)
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = style['border']
        cell.fill = style['fill']
        cell.number_format = style['number_format']


def clear_subject_rows(ws, start_row: int, end_row: int):
    """Clear cell values in subject area."""
    for row in range(start_row, end_row + 1):
        for col in range(1, 9):  # A through H
            ws.cell(row, col).value = None


def write_subjects(ws, subjects: list, start_row: int, ref_styles: dict,
                   default_height: float = 30):
    """Write subjects into worksheet starting at start_row.
    
    subjects: list of dicts with 'value' and 'row_height'
    """
    for i, subj in enumerate(subjects):
        row = start_row + i
        
        # Column A: sequential number
        ws.cell(row, 1).value = i + 1
        
        # Column B: subject name
        ws.cell(row, 2).value = subj['value']
        
        # Apply formatting
        apply_style(ws, row, ref_styles)
        
        # Set row height
        h = subj.get('row_height') or default_height
        ws.row_dimensions[row].height = h


def process_it_template(filepath: Path):
    """Redistribute subjects in an IT (F) template."""
    print(f"\nProcessing: {filepath.name}")
    backup_file(filepath)
    
    wb = openpyxl.load_workbook(filepath)
    sheets = wb.worksheets
    
    if len(sheets) != 4:
        print(f"  ERROR: Expected 4 sheets, found {len(sheets)}")
        return
    
    # 1. Collect all subjects
    all_subjects = collect_subjects(wb)
    total = len(all_subjects)
    expected_total = sum(IT_NEW_DIST)
    print(f"  Total subjects: {total} (expected: {expected_total})")
    
    if total != expected_total:
        print(f"  WARNING: subject count mismatch!")
        return
    
    # 2. Get reference styles from existing subject rows
    ref_style_p1 = get_reference_style(sheets[0], 15)  # Page 1 subject row
    ref_style_p2 = get_reference_style(sheets[1], 2)    # Page 2 subject row
    
    # 3. Split subjects according to new distribution
    splits = []
    idx = 0
    for count in IT_NEW_DIST:
        splits.append(all_subjects[idx:idx + count])
        idx += count
    
    # 4. Clear old subject areas and write new distribution
    
    # --- Page 1: subjects at rows 15..31 ---
    ws1 = sheets[0]
    old_max1 = ws1.max_row
    # Clear old subjects (rows 15..old_max)
    clear_subject_rows(ws1, 15, max(old_max1, 15 + IT_NEW_DIST[0]))
    # Write new subjects
    write_subjects(ws1, splits[0], 15, ref_style_p1)
    # Clear any leftover rows beyond new end
    for row in range(15 + IT_NEW_DIST[0], old_max1 + 1):
        for col in range(1, 9):
            ws1.cell(row, col).value = None
            # Remove borders from empty rows
            ws1.cell(row, col).border = Border()
    
    # --- Pages 2, 3, 4: subjects at rows 2..N ---
    for page_idx in range(1, 4):
        ws = sheets[page_idx]
        old_max = ws.max_row
        new_count = IT_NEW_DIST[page_idx]
        subj_list = splits[page_idx]
        
        # Clear old subjects
        clear_subject_rows(ws, 2, max(old_max, 2 + new_count))
        
        # Write new subjects
        write_subjects(ws, subj_list, 2, ref_style_p2)
        
        # Clear any leftover rows beyond new end
        for row in range(2 + new_count, old_max + 1):
            for col in range(1, 9):
                ws.cell(row, col).value = None
                ws.cell(row, col).border = Border()
    
    # 5. Save
    wb.save(filepath)
    wb.close()
    
    # 6. Verify
    wb2 = openpyxl.load_workbook(filepath)
    for ws in wb2.worksheets:
        count = sum(1 for r in range(1, ws.max_row + 1) 
                    if ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip())
        start_row = 15 if ws == wb2.worksheets[0] else 2
        first = None
        last = None
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v and str(v).strip():
                if first is None:
                    first = str(v)[:40]
                last = str(v)[:40]
        print(f"  {ws.title}: {count} subjects | {first} ... {last}")
    wb2.close()


def main():
    print("=" * 70)
    print("  Template Subject Redistribution")
    print("=" * 70)
    
    for tpath in IT_TEMPLATES:
        if not tpath.exists():
            print(f"\n  SKIP (not found): {tpath}")
            continue
        process_it_template(tpath)
    
    print("\n" + "=" * 70)
    print("  D (Accounting) templates: already balanced, no changes needed.")
    print("=" * 70)
    print("\nDone!")


if __name__ == "__main__":
    main()
