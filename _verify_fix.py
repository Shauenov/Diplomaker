"""Verify generated diploma data integrity against known issues."""
import openpyxl, os, sys

output_dir = r'c:\Users\user\OneDrive\Рабочий стол\template\output'

# Pick one IT file for testing
test_files = {
    'IT_KZ': None,
    'IT_RU': None,
    'ACC_KZ': None,
    'ACC_RU': None,
}

for fn in os.listdir(output_dir):
    if not fn.endswith('.xlsx'): continue
    if '3Ғ-2_Бейбіт' in fn and '_KZ' in fn: test_files['IT_KZ'] = fn
    if '3Ғ-2_Бейбіт' in fn and '_RU' in fn: test_files['IT_RU'] = fn
    if '3D-1_Асанов' in fn and '_KZ' in fn: test_files['ACC_KZ'] = fn
    if '3D-1_Асанов' in fn and '_RU' in fn: test_files['ACC_RU'] = fn

print("=== TEST FILES ===")
for k, v in test_files.items():
    print(f"  {k}: {v}")

def check_sheet(wb, sheet_idx, col_shift, checks, label):
    """Check specific rows for expected/unexpected values."""
    ws = wb.worksheets[sheet_idx]
    start_row = 15 if sheet_idx == 0 else 1
    
    # Build row mapping: entry_num -> actual_row
    entries = []
    for row in range(start_row, (ws.max_row or 200) + 1):
        v = ws.cell(row=row, column=2+col_shift).value
        if v and isinstance(v, str) and v.strip():
            entries.append((row, v.strip()[:60]))
    
    print(f"\n  --- {label} (sheet {sheet_idx}, {len(entries)} entries) ---")
    for entry_idx, expected_text, col_to_check, expected_val, description in checks:
        if entry_idx >= len(entries):
            print(f"  [SKIP] Entry #{entry_idx} out of range ({len(entries)} entries)")
            continue
        row, text = entries[entry_idx]
        actual = ws.cell(row=row, column=col_to_check+col_shift).value
        ok = True
        if expected_val is not None:
            ok = str(actual) == str(expected_val) if actual else expected_val == ''
        status = "OK" if ok else "FAIL"
        print(f"  [{status}] Entry #{entry_idx} '{text[:40]}': col{col_to_check}={actual} ({description})")

# IT Checks
for key, lang in [('IT_KZ', 'kz'), ('IT_RU', 'ru')]:
    fn = test_files[key]
    if not fn:
        print(f"\n[SKIP] {key} - no file found")
        continue
    fp = os.path.join(output_dir, fn)
    wb = openpyxl.load_workbook(fp)
    col_shift = 0 if key.endswith('KZ') else 0
    print(f"\n=== {key}: {fn} ===")
    
    # Page 1: Check БМ 1-4 have hours/credits (entries 13-16)
    check_sheet(wb, 0, 0, [
        (13, 'БМ 1', 3, '216', 'БМ 1 hours should be 216'),
        (14, 'БМ 2', 3, '72', 'БМ 2 hours should be 72'),
        (15, 'БМ 3', 3, '72', 'БМ 3 hours should be 72'),
        (16, 'БМ 4', 3, '24', 'БМ 4 hours should be 24'),
    ], 'Page 1 - БМ entries')
    
    # Page 2: Check ОН 1.1-1.4 have correct hours (entries 1-4, skipping КМ1 header at 0)
    cs2 = 1  # col_shift for sheet 1
    check_sheet(wb, 1, cs2, [
        (1, 'ОН 1.1', 3, '72', 'ОН 1.1 hours should be 72'),
        (2, 'ОН 1.2', 3, '48', 'ОН 1.2 hours should be 48'),
        (3, 'ОН 1.3', 3, '72', 'ОН 1.3 hours should be 72'),
        (4, 'ОН 1.4', 3, '72', 'ОН 1.4 hours should be 72'),
    ], 'Page 2 - ОН 1.x entries')
    
    # Page 4: Check КМ09 is at entry 0, ОН10.1 is NOT 504, Practice is at entry 10
    cs4 = 1  # col_shift for sheet 3
    check_sheet(wb, 3, cs4, [
        (0, 'КМ 09', 3, None, 'КМ 09 header at entry 0'),
        (5, 'ОН 10.1', 3, '24', 'ОН 10.1 hours should be 24, NOT 504'),
        (6, 'ОН 10.2', 3, '48', 'ОН 10.2 hours should be 48'),
        (7, 'ОН 10.3', 3, '72', 'ОН 10.3 hours should be 72'),
        (8, 'ОН 10.4', 3, '72', 'ОН 10.4 hours should be 72'),
        (9, 'Кәсіптік практика', 3, '504', 'Practice hours should be 504'),
    ], 'Page 4 - КМ09/10 entries')
    
    # Check ОН 10.3 does NOT have "сынақ" or "зачтено"
    ws4 = wb.worksheets[3]
    entries4 = []
    for row in range(1, (ws4.max_row or 200) + 1):
        v = ws4.cell(row=row, column=2+cs4).value
        if v and isinstance(v, str) and v.strip():
            entries4.append((row, v.strip()[:60]))
    if len(entries4) > 7:
        r10_3 = entries4[7][0]
        trad_val = ws4.cell(row=r10_3, column=8+cs4).value
        is_bad = trad_val and ('сынақ' in str(trad_val).lower() or 'зачтено' in str(trad_val).lower())
        print(f"  [{'FAIL' if is_bad else 'OK'}] ОН 10.3 traditional grade = '{trad_val}' (should NOT be сынақ/зачтено)")
    
    wb.close()

print("\n=== VERIFICATION COMPLETE ===")
