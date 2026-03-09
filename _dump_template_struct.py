import openpyxl, os
tdir = r'c:\Users\user\OneDrive\Рабочий стол\template\templates'
for fn in sorted(os.listdir(tdir)):
    if fn.endswith('.xlsx') and not fn.endswith('.bak'):
        fp = os.path.join(tdir, fn)
        wb = openpyxl.load_workbook(fp, read_only=True)
        print(f'=== {fn} ({len(wb.sheetnames)} sheets: {wb.sheetnames}) ===')
        for si, ws in enumerate(wb.worksheets):
            col_shift = 1 if si in (1,3) else 0
            start_row = 15 if si == 0 else 1
            entries = []
            max_r = ws.max_row or 200
            for row in range(start_row, max_r+1):
                v = ws.cell(row=row, column=2+col_shift).value
                if v and isinstance(v, str) and v.strip():
                    entries.append((row, v.strip()[:80]))
            print(f'  Sheet {si} "{ws.title}": {len(entries)} entries, rows {start_row}-{ws.max_row}')
            for r, t in entries:
                print(f'    row {r}: {t}')
        wb.close()
