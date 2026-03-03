import openpyxl

templates = [
    ('templates/Diplom_F_KZ_Template (4).xlsx', 'F_KZ'),
    ('templates/Diplom_F_RU_Template (4).xlsx', 'F_RU'),
    ('templates/Diplom_D_KZ_Template(4).xlsx', 'D_KZ'),
    ('templates/Diplom_D_RU_Template(4).xlsx', 'D_RU'),
]

for t, label in templates:
    try:
        wb = openpyxl.load_workbook(t)
        ws = wb.worksheets[0]
        print(f"\n=== {label} | Sheet: {ws.title} | Rows: {ws.max_row} ===")
        for r in range(1, min(20, ws.max_row + 1)):
            h = ws.row_dimensions[r].height
            vals = []
            for c in range(1, 9):
                v = ws.cell(r, c).value
                if v is not None:
                    vals.append(f"{chr(64+c)}={repr(v)[:50]}")
            hstr = f"{h:.1f}" if h else "None"
            vstr = "  ".join(vals) if vals else "(empty row)"
            print(f"  R{r:2d} h={hstr:>6s} | {vstr}")
        merged = [str(m) for m in ws.merged_cells.ranges]
        print(f"  Merged: {merged}")
        wb.close()
    except Exception as e:
        print(f"ERROR {label}: {e}")
