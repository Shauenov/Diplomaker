import os
import argparse
import pandas as pd
from typing import List

from configs import get_config
from src.parser import parse_excel_sheet
from src.generator import DiplomaGenerator
import sys

# Force utf-8 for Windows console
if sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def main():
    parser = argparse.ArgumentParser(description="Diploma Generator (Modular)")
    parser.add_argument("--source", type=str, required=True, help="Путь к исходному Excel-файлу")
    parser.add_argument("--group", type=str, required=True, choices=["3F", "3D"], help="Группа: 3F (IT) или 3D (Бухгалтеры)")
    parser.add_argument("--lang", type=str, default="ALL", choices=["KZ", "RU", "ALL"], help="Язык диплома")
    args = parser.parse_args()
    
    source_file = args.source
    if not os.path.exists(source_file):
        print(f"File not found: {source_file}")
        return
        
    os.makedirs("output", exist_ok=True)
    
    # 1. Загружаем Excel
    print(f"Loading '{source_file}'...", flush=True)
    import openpyxl
    wb = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
    
    target_prefix = args.group
    if target_prefix == "3F":
        target_prefix = "3Ғ"  # Автозамена на казахскую 'Ғ'
        
    target_sheets = [s for s in wb.sheetnames if s.startswith(target_prefix)]
    
    if not target_sheets:
        print(f"No sheets found starting with {target_prefix}")
        return
        
    langs_to_run = ["kz", "ru"] if args.lang == "ALL" else [args.lang.lower()]
    
    for sheet_name in target_sheets:
        print(f"\nProcessing sheet: {sheet_name}", flush=True)
        # Convert read-only worksheet to DataFrame safely (limit to 200x200 to prevent infinite row bug)
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(max_row=200, max_col=300, values_only=True):
            data.append(row)
        df = pd.DataFrame(data)
        students = parse_excel_sheet(df, sheet_name, start_row=5)
        print(f"  Found {len(students)} students.")
        
        for lang in langs_to_run:
            print(f"  Generating {lang.upper()} diplomas...")
            config, terms, template_name = get_config(args.group, lang)
            template_path = os.path.join("templates", template_name)
            
            if not os.path.exists(template_path):
                print(f"  [ERROR] Template {template_path} not found. Skipping {lang}.")
                continue
                
            for s in students:
                safe_name = s['name'].replace('/', ' ').replace('\\', ' ')
                out_name = f"{sheet_name}_{safe_name}_{lang.upper()}.xlsx"
                out_path = os.path.join("output", out_name)
                
                try:
                    print(f"    [DEBUG] Starting generator for {lang.upper()} template {template_path}...", flush=True)
                    generator = DiplomaGenerator(template_path, out_path, config, terms)
                    print(f"    [DEBUG] Fill student data...", flush=True)
                    generator.fill_student_data(s)
                    print(f"    [DEBUG] Closing...", flush=True)
                    generator.close()
                    print(f"    + {out_name}", flush=True)
                except Exception as e:
                    print(f"    - [ERROR] Failed to generate {out_name}: {e}")

if __name__ == "__main__":
    main()
