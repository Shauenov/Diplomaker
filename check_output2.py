import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

print("Checking 3D-1_Асанов Бекарыс Максатович_KZ.xlsx")
wb = openpyxl.load_workbook("output/3D-1_Асанов Бекарыс Максатович_KZ.xlsx")

for ws_name in ['Бет 1', 'Бет 2', 'Бет 3', 'Бет 4']:
    if ws_name not in wb.sheetnames:
        continue
    ws = wb[ws_name]
    print(f"\n=== {ws_name} ===")
    for row in range(1, 45):
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, str) and any(x in val for x in ['ПМ ', 'КМ ', 'ОН ', 'БМ ', 'Ф']):
            h = ws.cell(row=row, column=3).value
            c = ws.cell(row=row, column=4).value
            tr = ws.cell(row=row, column=8).value
            print(f"Row {row:02d}: {str(val)[:40]:<40} | H: {h} C: {c} Trad: {tr}")
