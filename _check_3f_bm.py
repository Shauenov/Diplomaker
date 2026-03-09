"""Quick verification of IT (3F) diploma БМ grades."""
import openpyxl, os

out_dir = "output"
files_3f = [f for f in os.listdir(out_dir) if f.startswith("3Ғ-") and f.endswith("_KZ.xlsx") and "nan" not in f]
fp = os.path.join(out_dir, files_3f[0])
print(f"Checking: {files_3f[0]}")

wb = openpyxl.load_workbook(fp)
ws = wb.worksheets[0]  # Sheet 0 — page 1 has БМ entries 14-17

for row in range(15, ws.max_row + 1):
    text = ws.cell(row=row, column=2).value
    if text and "БМ" in str(text):
        h = ws.cell(row=row, column=3).value
        c = ws.cell(row=row, column=4).value
        p = ws.cell(row=row, column=5).value
        l = ws.cell(row=row, column=6).value
        g = ws.cell(row=row, column=7).value
        t = ws.cell(row=row, column=8).value
        has_grade = p is not None or l is not None or t is not None
        status = "OK" if has_grade else "MISSING"
        print(f"  [{status}] {str(text)[:50]:50s} h={h} c={c} p={p} l={l} g={g} t={t}")

wb.close()
