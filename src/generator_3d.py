# -*- coding: utf-8 -*-
"""
3D (Accounting) Diploma Generator — Hardcoded Template Filler
==============================================================
Специализированный генератор для нового 2-листового шаблона дипломов
бухгалтерской специальности (3D groups).

Шаблон: Diplom_D_RU_Template(4).xlsx  /  Diplom_D_KZ_Template(4).xlsx
Структура: 2 листа («Лист 1», «Лист 2»)
  Лист 1  — левая часть (items 1-17, cols A-H)
            правая часть (items 18-29, cols K-R)
  Лист 2  — левая часть (items 30-42, cols A-H)
            правая часть (items 43-50, cols K-R)

Items 45-50 записываются через merged-блок с \\n-разделителями.

Скрипт НЕ перезаписывает названия предметов (subject_name), т.к.
в шаблоне они в сложных merged cells. Пишутся только оценки/часы/кредиты.
"""

import gc
import re
from typing import Dict, Any, Optional
from zipfile import ZIP_DEFLATED, ZipFile
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import ExcelWriter

from .page1_alignment import get_page1_header_alignment, write_aligned_field


# ═════════════════════════════════════════════════════════════════
# CELL MAP — Жёсткий маппинг координат для каждого пункта диплома
# ═════════════════════════════════════════════════════════════════

_LEFT_COLS = {
    "hours": 3,        # C
    "credits": 4,      # D
    "score_100": 5,    # E
    "grade_letter": 6, # F
    "gpa_numeric": 7,  # G
    "final_grade": 8,  # H
}

_RIGHT_COLS = {
    "hours": 13,       # M
    "credits": 14,     # N
    "score_100": 15,   # O
    "grade_letter": 16,# P
    "gpa_numeric": 17, # Q
    "final_grade": 18, # R
}

# Левая часть 1 листа одинакова для RU/KZ (15-31)
_COMMON_SHEET1_LEFT = {
    i: {"sheet": 0, "row": 14 + i, "cols": _LEFT_COLS, "is_header": False}
    for i in range(1, 18)
}

# ----------------- RU MAPPING -----------------
CELL_MAP_3D_RU: Dict[int, Dict[str, Any]] = {
    **_COMMON_SHEET1_LEFT,
    
    # Лист 1 Правая часть
    # 18: skip (is_header)
    19: {"sheet": 0, "row": 3,  "cols": _RIGHT_COLS, "is_header": False},
    20: {"sheet": 0, "row": 5,  "cols": _RIGHT_COLS, "is_header": False},
    21: {"sheet": 0, "row": 9,  "cols": _RIGHT_COLS, "is_header": False},
    22: {"sheet": 0, "row": 12, "cols": _RIGHT_COLS, "is_header": False},
    # 23: skip
    24: {"sheet": 0, "row": 16, "cols": _RIGHT_COLS, "is_header": False},
    25: {"sheet": 0, "row": 19, "cols": _RIGHT_COLS, "is_header": False},
    26: {"sheet": 0, "row": 22, "cols": _RIGHT_COLS, "is_header": False},
    27: {"sheet": 0, "row": 25, "cols": _RIGHT_COLS, "is_header": False},
    # 28: skip
    29: {"sheet": 0, "row": 29, "cols": _RIGHT_COLS, "is_header": False},

    # Лист 2 Левая часть (30-42)
    **{
        i: {"sheet": 1, "row": i - 29, "cols": _LEFT_COLS, "is_header": i in (33, 41)}
        for i in range(30, 43)
    },

    # Лист 2 Правая часть
    43: {"sheet": 1, "row": 1, "cols": _RIGHT_COLS, "is_header": False},
    44: {"sheet": 1, "row": 2, "cols": _RIGHT_COLS, "is_header": False},
}


# ----------------- KZ MAPPING -----------------
CELL_MAP_3D_KZ: Dict[int, Dict[str, Any]] = {
    **_COMMON_SHEET1_LEFT,
    
    # Лист 1 Правая часть
    # 18: skip
    19: {"sheet": 0, "row": 3,  "cols": _RIGHT_COLS, "is_header": False},
    20: {"sheet": 0, "row": 5,  "cols": _RIGHT_COLS, "is_header": False},
    21: {"sheet": 0, "row": 8,  "cols": _RIGHT_COLS, "is_header": False},
    22: {"sheet": 0, "row": 11, "cols": _RIGHT_COLS, "is_header": False},
    # 23: skip
    24: {"sheet": 0, "row": 15, "cols": _RIGHT_COLS, "is_header": False},
    25: {"sheet": 0, "row": 18, "cols": _RIGHT_COLS, "is_header": False},
    26: {"sheet": 0, "row": 21, "cols": _RIGHT_COLS, "is_header": False},
    27: {"sheet": 0, "row": 22, "cols": _RIGHT_COLS, "is_header": False},
    # 28: skip
    29: {"sheet": 0, "row": 27, "cols": _RIGHT_COLS, "is_header": False},

    # Лист 2 Левая часть (30-41)
    **{
        i: {"sheet": 1, "row": i - 29, "cols": _LEFT_COLS, "is_header": i in (33, 41)}
        for i in range(30, 42)
    },

    # Лист 2 Правая часть
    42: {"sheet": 1, "row": 1, "cols": _RIGHT_COLS, "is_header": False},
    43: {"sheet": 1, "row": 2, "cols": _RIGHT_COLS, "is_header": False},
}

# ═════════════════════════════════════════════════════════════════
# Индексы для merged блока 45-50 (RU) и 44-50 (KZ)
# ═════════════════════════════════════════════════════════════════
_BLOCK_LINE_INDICES_RU = {
    45: 0,   # Итоговая аттестация
    46: 2,   # line 3
    47: 3,   # line 4
    48: 4,   # line 5
    49: 7,   # line 8
    50: 12,  # line 13
}

_BLOCK_LINE_INDICES_KZ = {
    44: 1,   # Оценка практики (после 1 Enter)
    45: 7,   # Оценка аттестации (после 5 Enter от практики)
    46: 9,   # Факультатив 1 (после 1 Enter от аттестации)
    47: 10,  # Факультатив 2
    48: 11,  # Факультатив 3
    49: 14,  # Факультатив 4 (после 2 Enter от Ф3)
    50: 17,  # Факультатив 5 (после 2 Enter от Ф4)
}

_BLOCK_TOTAL_LINES_KZ = 18
_BLOCK_TOTAL_LINES = 13


# ═════════════════════════════════════════════════════════════════
# Шапка диплома — координаты на Лист 1
# ═════════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════════
# Шапка диплома — координаты на Лист 1
# ═════════════════════════════════════════════════════════════════
#
# (row, col) — 1-индексированые (как в openpyxl)
# Col числа: A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8
#
# ВАЖНО: Google Sheets НЕ отображает Excel indent для merged cells.
# Поэтому используем horizontal='center' — как в эталонном файле (Гаухар).
#
# Merged cells в шапке:
#   C2:D2  — номер диплома
#   B3:H3  — ФИО
#   B4:C4  — год поступления
#   B5:H5  — колледж
#   B6:H6  — специальность (только RU; KZ без merge)
#   B9:H9  — квалификация


# ═════════════════════════════════════════════════════════════════
# Генератор
# ═════════════════════════════════════════════════════════════════

class DiplomaGenerator3D:
    """
    Генератор дипломов для 3D (Бухгалтерия) — новый 2-листовой шаблон.

    Использует жёсткий маппинг координат (CELL_MAP_3D) вместо
    построчного сканирования шаблона.

    Usage:
        gen = DiplomaGenerator3D(template_path, output_path, terms)
        gen.fill_student_data(student_dict, structured_pages, lang='ru')
        gen.close()
    """

    def __init__(
        self,
        template_path: str,
        output_path: str,
        terms: Dict[str, str],
    ):
        self.template_path = template_path
        self.output_path = output_path
        self.terms = terms
        self.workbook = openpyxl.load_workbook(template_path)

    # ─────────────────────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────────────────────

    def fill_student_data(
        self,
        student: Dict[str, Any],
        structured_pages: Dict[int, list],
        lang: str = "ru",
    ):
        """
        Заполняет диплом данными студента.

        Args:
            student: dict с ключами 'name', 'diploma_num', 'year_start',
                     'year_end', 'meta', 'grades'
            structured_pages: {page_num: [entry, ...]} из core.bridge
            lang: 'kz' или 'ru'
        """
        is_kz = lang.lower() == "kz"

        # 1. Заполняем шапку
        ws0 = self.workbook.worksheets[0]
        self._fill_header(ws0, student, is_kz)

        # 2. Собираем плоский список всех entries (page 1→2→3→4)
        all_entries = []
        for page_num in sorted(structured_pages.keys()):
            all_entries.extend(structured_pages[page_num])

        # 3. Заполняем изолированные items 1...N
        max_item = 43 if is_kz else 44
        cell_map = CELL_MAP_3D_KZ if is_kz else CELL_MAP_3D_RU

        for item_num, entry in enumerate(all_entries, start=1):
            if item_num > max_item:
                break  # 나머지는 블록으로

            cell_info = cell_map.get(item_num)
            if not cell_info:
                continue

            ws = self.workbook.worksheets[cell_info["sheet"]]
            row = cell_info["row"]
            cols = cell_info["cols"]
            is_header = cell_info.get("is_header", False)
            
            # МОДУЛИ ПУСТУЮТ: никаких часов, кредитов, оценок
            if is_header:
                continue

            hours = entry.get("hours", "")
            credits_val = entry.get("credits", "")

            # Пишем часы и кредиты
            if hours:
                self._write(ws, row, cols["hours"], hours)
            if credits_val:
                self._write(ws, row, cols["credits"], credits_val)

            # Пишем оценки
            self._write_grades(ws, row, cols, entry, is_kz)

        # 4. Заполняем блок (merged cells на Лист 2)
        # Блок начинается с 44 (для KZ) или 45 (для RU)
        block_start_index = max_item  # 43 or 44 (0-indexed base element)
        block_entries = all_entries[block_start_index:]
        if block_entries:
            self._fill_block_45_50(block_entries, is_kz)

        print(f"[3D Generator] Filled: {student.get('name', '???')}")

    def close(self):
        """Сохраняет файл и очищает память."""
        if self.workbook is None:
            return

        try:
            self._save_workbook_fast(self.output_path)
        finally:
            self.workbook.close()
            self.workbook = None
            gc.collect()

    def _save_workbook_fast(self, output_path: str) -> None:
        with ZipFile(output_path, "w", ZIP_DEFLATED, allowZip64=True, compresslevel=1) as archive:
            ExcelWriter(self.workbook, archive).save()

    # ─────────────────────────────────────────────────────────
    # Шапка
    # ─────────────────────────────────────────────────────────

    def _fill_header(self, ws, student: Dict[str, Any], is_kz: bool):
        """
        Заполняет шапку на Лист 1.

        Выравнивание через indent:
          - INDENT_HALF (4) ≈ B(half)  — ФИО, колледж
          - INDENT_3_4  (6) ≈ B(3/4)  — специальность, квалификация
        
        """
        meta = student.get("meta", {})
        header_cfg = get_page1_header_alignment(is_kz)

        # Номер диплома (C2, merged C2:D2)
        diploma_num = student.get("diploma_num", "")
        if diploma_num and str(diploma_num) not in ("nan", ""):
            write_aligned_field(ws, header_cfg["diplom_id"], diploma_num)

        # ФИО
        full_name = student.get("name", "")
        if full_name:
            write_aligned_field(ws, header_cfg["full_name"], full_name)

        # Год поступления (B4, merged B4:C4)
        year_start = student.get("year_start", "")
        if year_start:
            write_aligned_field(ws, header_cfg["start_year"], year_start)

        # Год окончания (F4, не в merge)
        year_end = student.get("year_end", "")
        if year_end:
            write_aligned_field(ws, header_cfg["end_year"], year_end)

        # Колледж (B5, merged B5:H5)
        college = (
            "Жамбыл инновациялық жоғары колледжінде"
            if is_kz
            else "Жамбылском инновационным высшем колледже"
        )
        write_aligned_field(ws, header_cfg["college"], college)

        # Специальность (B6, merged B6:H6)
        specialty = meta.get("specialty_kz" if is_kz else "specialty_ru", "")
        if specialty:
            write_aligned_field(ws, header_cfg["speciality"], specialty)

        # Квалификация (B9, merged B9:H9)
        qualification = meta.get(
            "qualification_kz" if is_kz else "qualification_ru", ""
        )
        if qualification:
            write_aligned_field(ws, header_cfg["qualification"], qualification)

    # ─────────────────────────────────────────────────────────
    # Запись оценок для одного предмета (items 1-44)
    # ─────────────────────────────────────────────────────────

    def _write_grades(
        self,
        ws,
        row: int,
        cols: dict,
        entry: dict,
        is_kz: bool,
    ):
        """Записывает оценки одного предмета по координатам."""
        subj_obj = entry.get("subject")
        is_elective = subj_obj.is_elective if subj_obj else False
        is_practice = entry.get("is_practice", False)

        trad = entry.get("traditional_kz" if is_kz else "traditional_ru", "")
        pts = entry.get("points", "")
        let = entry.get("letter", "")
        gpa = entry.get("gpa", "")

        # Факультативы → зачтено / сынақ
        if is_elective:
            trad = self.terms.get("traditional_elective", "зачтено")
            # Для факультативов не пишем баллы / букву / GPA
            pts, let, gpa = "", "", ""

        # Практика без оценки → зачтено
        elif is_practice and not trad:
            hours = entry.get("hours", "")
            credits_val = entry.get("credits", "")
            if not hours and not credits_val:
                trad = self.terms.get("traditional_practice", "зачтено")

        if pts:
            self._write(ws, row, cols["score_100"], pts)
        if let:
            self._write(ws, row, cols["grade_letter"], let)
        if gpa:
            self._write(ws, row, cols["gpa_numeric"], gpa)
        if trad:
            self._write(ws, row, cols["final_grade"], trad, is_trad=True)

    # ─────────────────────────────────────────────────────────
    # Merged-блок items 45–50 (Лист 2, правая часть, row 3+)
    # ─────────────────────────────────────────────────────────

    def _fill_block_45_50(self, entries: list, is_kz: bool):
        """
        Заполняет merged-блок на Лист 2 (row 3+).
        """
        ws = self.workbook.worksheets[1]
        block_row = 3

        total_lines = _BLOCK_TOTAL_LINES_KZ if is_kz else _BLOCK_TOTAL_LINES
        hours_lines = [""] * total_lines
        credits_lines = [""] * total_lines
        score_lines = [""] * total_lines
        letter_lines = [""] * total_lines
        gpa_lines = [""] * total_lines
        grade_lines = [""] * total_lines

        # Для хардкода часов и кредитов факультативов
        elective_hours_credits = {
            46: ("120", "5"),
            47: ("72", "3"),
            48: ("48", "2"),
            49: ("36", "1,5"),
            50: ("24", "1"),
        }

        block_indices = _BLOCK_LINE_INDICES_KZ if is_kz else _BLOCK_LINE_INDICES_RU
        base_item = 44 if is_kz else 45

        for i, entry in enumerate(entries):
            item_num = base_item + i
            if item_num > 50:
                break

            line_idx = block_indices.get(item_num)
            if line_idx is None:
                continue

            subj_obj = entry.get("subject")
            is_elective = subj_obj.is_elective if subj_obj else False
            is_practice = entry.get("is_practice", False)

            hours = entry.get("hours", "")
            credits_val = entry.get("credits", "")
            trad = entry.get("traditional_kz" if is_kz else "traditional_ru", "")
            points = entry.get("points", "")
            letter = entry.get("letter", "")
            gpa = entry.get("gpa", "")

            # Факультативы (хардкод часов/кредитов)
            if item_num in elective_hours_credits:
                trad = self.terms.get("traditional_elective", "зачтено")
                h, c = elective_hours_credits[item_num]
                hours = h
                credits_val = c
            elif is_practice:
                if not trad:
                    trad = self.terms.get("traditional_practice", "зачтено")

            if hours:
                hours_lines[line_idx] = str(hours)
            if credits_val:
                credits_lines[line_idx] = str(credits_val)
            if points:
                score_lines[line_idx] = str(points)
            if letter:
                letter_lines[line_idx] = str(letter)
            if gpa:
                gpa_lines[line_idx] = str(gpa)
            if trad:
                trad_line_idx = line_idx
                if trad_line_idx >= 0:
                    grade_lines[trad_line_idx] = str(trad)

        # Записываем merged-ячейки
        # Записываем все колонки через \n.join
        # Теперь O, P, Q будут иметь те же отступы, что и R
        ws.cell(row=block_row, column=13).value = "\n".join(hours_lines)    # M
        ws.cell(row=block_row, column=14).value = "\n".join(credits_lines)  # N
        ws.cell(row=block_row, column=15).value = "\n".join(score_lines)    # O
        ws.cell(row=block_row, column=16).value = "\n".join(letter_lines)   # P
        ws.cell(row=block_row, column=17).value = "\n".join(gpa_lines)      # Q
        ws.cell(row=block_row, column=18).value = "\n".join(grade_lines)  # R

    # ─────────────────────────────────────────────────────────
    # Агрегация часов модулей
    # ─────────────────────────────────────────────────────────

    def _aggregate_module_hours(
        self,
        header_entry: dict,
        structured_pages: Dict[int, list],
    ) -> tuple:
        """
        Подсчитывает сумму часов/кредитов подчинённых ОН для
        заголовка модуля (КМ/ПМ).
        """
        from .utils import normalize_key

        subj_obj = header_entry.get("subject")
        if not subj_obj:
            return "", ""

        m = re.search(
            r"(КМ|ПМ|БМ|ОН)\s*0*(\d+)",
            subj_obj.name_kz,
            re.IGNORECASE,
        )
        if not m:
            return "", ""

        mod_type = m.group(1).lower()
        if mod_type == "бм":
            return "", ""  # БМ часы не агрегируются

        mod_num = m.group(2)
        prefix_search = (
            f"он{mod_num}"
            if mod_type in ("км", "пм")
            else f"{mod_type}{mod_num}"
        )

        th, tc = 0.0, 0.0
        for pe_list in structured_pages.values():
            for pe in pe_list:
                nk = normalize_key(pe["subject"].name_kz)
                if nk.startswith(prefix_search):
                    h_str = str(pe.get("hours", "0")).replace(",", ".")
                    c_str = str(pe.get("credits", "0")).replace(",", ".")
                    try:
                        th += float(h_str) if h_str.replace(".", "", 1).isdigit() else 0
                    except ValueError:
                        pass
                    try:
                        tc += float(c_str) if c_str.replace(".", "", 1).isdigit() else 0
                    except ValueError:
                        pass

        hours = str(int(th)) if th > 0 and th == int(th) else (str(th) if th > 0 else "")
        credits_val = str(int(tc)) if tc > 0 and tc == int(tc) else (str(tc) if tc > 0 else "")
        return hours, credits_val

    # ─────────────────────────────────────────────────────────
    # Утилита записи
    # ─────────────────────────────────────────────────────────

    def _write(self, ws, row: int, col: int, val, is_trad: bool = False):
        """Записывает значение в ячейку, сохраняя alignment шаблона."""
        cell = ws.cell(row=row, column=col)
        cell.value = val
