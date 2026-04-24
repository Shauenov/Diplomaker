# -*- coding: utf-8 -*-
"""
IT (3F) Diploma Generator — Hardcoded Template Filler
=====================================================
Специализированный генератор для нового 2-листового шаблона дипломов
IT специальности (3F groups).

Шаблон: Diplom_F_RU_Template(4).xlsx  /  Diplom_F_KZ_Template(4).xlsx
Структура: 2 листа («Лист 1», «Лист 2»)
  Лист 1  — левая часть (items 1-17, cols A-H, rows 15-31)
            правая часть (items 18-34, cols K-R)
  Лист 2  — левая часть (RU: items 35-52; KZ: items 35-53, cols A-H)
            правая часть (RU: items 53-65; KZ: items 54-65, cols K-R)

RU items 61-65 записываются через merged-блок с \\n-разделителями.
KZ items 61-65 записываются в отдельные ячейки.

Скрипт НЕ перезаписывает названия предметов, т.к.
в шаблоне они в сложных merged cells. Пишутся только оценки/часы/кредиты.
"""

import gc
import re
from typing import Dict, Any
from zipfile import ZIP_DEFLATED, ZipFile
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import ExcelWriter

from .page1_alignment import get_page1_header_alignment, write_aligned_field


# ═════════════════════════════════════════════════════════════════
# CELL MAP — Жёсткий маппинг координат для каждого пункта диплома
# ═════════════════════════════════════════════════════════════════

_LEFT_COLS = {
    "hours": 3,        # C
    "credits": 4,      # D
    "score_100": 5,    # E
    "grade_letter": 6, # F
    "gpa_numeric": 7,  # G
    "final_grade": 8,  # H
}

_RIGHT_COLS = {
    "hours": 13,       # M
    "credits": 14,     # N
    "score_100": 15,   # O
    "grade_letter": 16,# P
    "gpa_numeric": 17, # Q
    "final_grade": 18, # R
}

# Левая часть 1-го листа одинакова для RU/KZ (items 1-17, rows 15-31)
_COMMON_SHEET1_LEFT = {
    i: {"sheet": 0, "row": 14 + i, "cols": _LEFT_COLS, "is_header": False}
    for i in range(1, 18)
}

# MODULE headers (не записываем оценки): items 18, 23, 27, 31, 35, 39, 44, 48, 52, 56
_MODULE_ITEMS = {18, 23, 27, 31, 35, 39, 44, 48, 52, 56}


# ----------------- RU MAPPING -----------------
CELL_MAP_IT_RU: Dict[int, Dict[str, Any]] = {
    **_COMMON_SHEET1_LEFT,

    # Лист 1 Правая часть (items 18-34)
    18: {"sheet": 0, "row": 1,  "cols": _RIGHT_COLS, "is_header": True},   # ПМ01
    19: {"sheet": 0, "row": 2,  "cols": _RIGHT_COLS, "is_header": False},
    20: {"sheet": 0, "row": 3,  "cols": _RIGHT_COLS, "is_header": False},
    21: {"sheet": 0, "row": 5,  "cols": _RIGHT_COLS, "is_header": False},
    22: {"sheet": 0, "row": 6,  "cols": _RIGHT_COLS, "is_header": False},
    23: {"sheet": 0, "row": 8,  "cols": _RIGHT_COLS, "is_header": True},   # ПМ02
    24: {"sheet": 0, "row": 10, "cols": _RIGHT_COLS, "is_header": False},
    25: {"sheet": 0, "row": 11, "cols": _RIGHT_COLS, "is_header": False},
    26: {"sheet": 0, "row": 12, "cols": _RIGHT_COLS, "is_header": False},
    27: {"sheet": 0, "row": 14, "cols": _RIGHT_COLS, "is_header": True},   # ПМ03
    28: {"sheet": 0, "row": 15, "cols": _RIGHT_COLS, "is_header": False},
    29: {"sheet": 0, "row": 17, "cols": _RIGHT_COLS, "is_header": False},
    30: {"sheet": 0, "row": 20, "cols": _RIGHT_COLS, "is_header": False},
    31: {"sheet": 0, "row": 22, "cols": _RIGHT_COLS, "is_header": True},   # ПМ04
    32: {"sheet": 0, "row": 24, "cols": _RIGHT_COLS, "is_header": False},
    33: {"sheet": 0, "row": 27, "cols": _RIGHT_COLS, "is_header": False},
    34: {"sheet": 0, "row": 28, "cols": _RIGHT_COLS, "is_header": False},

    # Лист 2 Левая часть (items 35-52, rows 1-18)
    **{
        i: {"sheet": 1, "row": i - 34, "cols": _LEFT_COLS, "is_header": i in (35, 39, 44, 48, 52)}
        for i in range(35, 53)
    },

    # Лист 2 Правая часть (items 53-60, rows 1-8)
    53: {"sheet": 1, "row": 1, "cols": _RIGHT_COLS, "is_header": False},
    54: {"sheet": 1, "row": 2, "cols": _RIGHT_COLS, "is_header": False},
    55: {"sheet": 1, "row": 3, "cols": _RIGHT_COLS, "is_header": False},
    56: {"sheet": 1, "row": 4, "cols": _RIGHT_COLS, "is_header": True},   # ПМ10
    57: {"sheet": 1, "row": 5, "cols": _RIGHT_COLS, "is_header": False},
    58: {"sheet": 1, "row": 6, "cols": _RIGHT_COLS, "is_header": False},
    59: {"sheet": 1, "row": 7, "cols": _RIGHT_COLS, "is_header": False},
    60: {"sheet": 1, "row": 8, "cols": _RIGHT_COLS, "is_header": False},
    # Items 61-65 → merged block (handled by _fill_block_ru)
}

# ----------------- KZ MAPPING -----------------
CELL_MAP_IT_KZ: Dict[int, Dict[str, Any]] = {
    **_COMMON_SHEET1_LEFT,

    # Лист 1 Правая часть (items 18-34)
    18: {"sheet": 0, "row": 2,  "cols": _RIGHT_COLS, "is_header": True},   # КМ01
    19: {"sheet": 0, "row": 3,  "cols": _RIGHT_COLS, "is_header": False},
    20: {"sheet": 0, "row": 5,  "cols": _RIGHT_COLS, "is_header": False},
    21: {"sheet": 0, "row": 6,  "cols": _RIGHT_COLS, "is_header": False},
    22: {"sheet": 0, "row": 8,  "cols": _RIGHT_COLS, "is_header": False},
    23: {"sheet": 0, "row": 9,  "cols": _RIGHT_COLS, "is_header": True},   # КМ02
    24: {"sheet": 0, "row": 11, "cols": _RIGHT_COLS, "is_header": False},
    25: {"sheet": 0, "row": 12, "cols": _RIGHT_COLS, "is_header": False},
    26: {"sheet": 0, "row": 14, "cols": _RIGHT_COLS, "is_header": False},
    27: {"sheet": 0, "row": 15, "cols": _RIGHT_COLS, "is_header": True},   # КМ03
    28: {"sheet": 0, "row": 17, "cols": _RIGHT_COLS, "is_header": False},
    29: {"sheet": 0, "row": 19, "cols": _RIGHT_COLS, "is_header": False},
    30: {"sheet": 0, "row": 22, "cols": _RIGHT_COLS, "is_header": False},
    31: {"sheet": 0, "row": 23, "cols": _RIGHT_COLS, "is_header": True},   # КМ04
    32: {"sheet": 0, "row": 27, "cols": _RIGHT_COLS, "is_header": False},
    33: {"sheet": 0, "row": 29, "cols": _RIGHT_COLS, "is_header": False},
    34: {"sheet": 0, "row": 30, "cols": _RIGHT_COLS, "is_header": False},

    # Лист 2 Левая часть (items 35-53, rows 1-19) — KZ имеет на 1 item больше, чем RU
    **{
        i: {"sheet": 1, "row": i - 34, "cols": _LEFT_COLS, "is_header": i in (35, 39, 44, 48, 52)}
        for i in range(35, 54)
    },

    # Лист 2 Правая часть (items 54-65, rows 1-12) — все индивидуальные ячейки
    54: {"sheet": 1, "row": 1,  "cols": _RIGHT_COLS, "is_header": False},
    55: {"sheet": 1, "row": 2,  "cols": _RIGHT_COLS, "is_header": False},
    56: {"sheet": 1, "row": 3,  "cols": _RIGHT_COLS, "is_header": True},   # КМ10
    57: {"sheet": 1, "row": 4,  "cols": _RIGHT_COLS, "is_header": False},
    58: {"sheet": 1, "row": 5,  "cols": _RIGHT_COLS, "is_header": False},
    59: {"sheet": 1, "row": 6,  "cols": _RIGHT_COLS, "is_header": False},
    60: {"sheet": 1, "row": 7,  "cols": _RIGHT_COLS, "is_header": False},
    61: {"sheet": 1, "row": 8,  "cols": _RIGHT_COLS, "is_header": False},   # Practice
    62: {"sheet": 1, "row": 9,  "cols": _RIGHT_COLS, "is_header": False},   # Attestation
    63: {"sheet": 1, "row": 10, "cols": _RIGHT_COLS, "is_header": False},   # Elective
    64: {"sheet": 1, "row": 11, "cols": _RIGHT_COLS, "is_header": False},   # Elective
    65: {"sheet": 1, "row": 12, "cols": _RIGHT_COLS, "is_header": False},   # Elective
}


# ═════════════════════════════════════════════════════════════════
# Индексы для merged блока items 61-65 (RU only)
# Merged cells: O9:O11, P9:P11, Q9:Q11, R9:R11 на Лист 2
# ═════════════════════════════════════════════════════════════════
# Template R9 value: '\n\n\n\n\n\nзачтено\nзачтено\nзачтено'  (9 lines total)
_BLOCK_LINE_INDICES_IT_RU = {
    61: 0,   # Practice traditional
    62: 5,   # Attestation traditional (shifted down by 4 lines)
    63: 6,   # Elective 1
    64: 7,   # Elective 2
    65: 8,   # Elective 3
}
_BLOCK_TOTAL_LINES_IT = 9


# ═════════════════════════════════════════════════════════════════
# Генератор
# ═════════════════════════════════════════════════════════════════

class DiplomaGeneratorIT:
    """
    Генератор дипломов для 3F (IT) — новый 2-листовой шаблон.

    Использует жёсткий маппинг координат (CELL_MAP_IT) вместо
    построчного сканирования шаблона.

    Usage:
        gen = DiplomaGeneratorIT(template_path, output_path, terms)
        gen.fill_student_data(student_dict, structured_pages, lang='ru')
        gen.close()
    """

    def __init__(
        self,
        template_path: str,
        output_path: str,
        terms: Dict[str, str],
    ):
        self.template_path = template_path
        self.output_path = output_path
        self.terms = terms
        self.workbook = openpyxl.load_workbook(template_path)

    # ─────────────────────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────────────────────

    def fill_student_data(
        self,
        student: Dict[str, Any],
        structured_pages: Dict[int, list],
        lang: str = "ru",
    ):
        """
        Заполняет диплом данными студента.

        Args:
            student: dict с ключами 'name', 'diploma_num', 'year_start',
                     'year_end', 'meta', 'grades'
            structured_pages: {page_num: [entry, ...]} из core.bridge
            lang: 'kz' или 'ru'
        """
        is_kz = lang.lower() == "kz"

        # 1. Заполняем шапку
        ws0 = self.workbook.worksheets[0]
        self._fill_header(ws0, student, is_kz)

        # 2. Собираем плоский список всех entries (page 1→2→3→4)
        all_entries = []
        for page_num in sorted(structured_pages.keys()):
            all_entries.extend(structured_pages[page_num])

        # 3. Определяем cell map и лимит для изолированных ячеек
        cell_map = CELL_MAP_IT_KZ if is_kz else CELL_MAP_IT_RU

        # Для RU: items 1-60 — в CELL_MAP, items 61-65 — в merged block
        # Для KZ: items 1-65 — все в CELL_MAP (включая 61-65)
        max_isolated_item = 65 if is_kz else 60

        for item_num, entry in enumerate(all_entries, start=1):
            if item_num > max_isolated_item:
                break

            cell_info = cell_map.get(item_num)
            if not cell_info:
                continue

            ws = self.workbook.worksheets[cell_info["sheet"]]
            row = cell_info["row"]
            cols = cell_info["cols"]
            is_header = cell_info.get("is_header", False)

            # MODULE headers — пропускаем оценки
            if is_header:
                continue

            hours = entry.get("hours", "")
            credits_val = entry.get("credits", "")

            # Пишем часы и кредиты
            if hours:
                self._write(ws, row, cols["hours"], hours)
            if credits_val:
                self._write(ws, row, cols["credits"], credits_val)

            # Пишем оценки
            self._write_grades(ws, row, cols, entry, is_kz)

        # 4. RU: Заполняем merged block (items 61-65 на Лист 2, row 9)
        if not is_kz:
            block_entries = all_entries[60:]  # items 61+ (0-indexed: 60)
            if block_entries:
                self._fill_block_ru(block_entries)

        print(f"[IT Generator] Filled: {student.get('name', '???')}")

    def close(self):
        """Сохраняет файл и очищает память."""
        if self.workbook is None:
            return

        try:
            self._save_workbook_fast(self.output_path)
        finally:
            self.workbook.close()
            self.workbook = None
            gc.collect()

    def _save_workbook_fast(self, output_path: str) -> None:
        with ZipFile(output_path, "w", ZIP_DEFLATED, allowZip64=True, compresslevel=1) as archive:
            ExcelWriter(self.workbook, archive).save()

    # ─────────────────────────────────────────────────────────
    # Шапка
    # ─────────────────────────────────────────────────────────

    def _fill_header(self, ws, student: Dict[str, Any], is_kz: bool):
        """
        Заполняет шапку в левой части Лист 1 (A-H).
        Выравнивание (alignment) не трогаем, за исключением ФИО и названий
        организаций, где применяем специальный indent для соответствия шаблону.
        
        """
        meta = student.get("meta", {})
        header_cfg = get_page1_header_alignment(is_kz)

        # Номер диплома (C2, merged C2:D2)
        diploma_num = student.get("diploma_num", "")
        if diploma_num and str(diploma_num) not in ("nan", ""):
            write_aligned_field(ws, header_cfg["diplom_id"], diploma_num)

        # ФИО
        full_name = student.get("name", "")
        if full_name:
            write_aligned_field(ws, header_cfg["full_name"], full_name)

        # Год поступления (B4, merged B4:C4)
        year_start = student.get("year_start", "")
        if year_start:
            write_aligned_field(ws, header_cfg["start_year"], year_start)

        # Год окончания (F4)
        year_end = student.get("year_end", "")
        if year_end:
            write_aligned_field(ws, header_cfg["end_year"], year_end)

        # Колледж (B5, merged B5:H5)
        college = (
            "Жамбыл инновациялық жоғары колледжінде"
            if is_kz
            else "Жамбылском инновационным высшем колледже"
        )
        write_aligned_field(ws, header_cfg["college"], college)

        # Специальность
        specialty = meta.get("specialty_kz" if is_kz else "specialty_ru", "")
        if specialty:
            write_aligned_field(ws, header_cfg["speciality"], specialty)

        # Квалификация (B9, merged B9:H9)
        qualification = meta.get(
            "qualification_kz" if is_kz else "qualification_ru", ""
        )
        if qualification:
            write_aligned_field(ws, header_cfg["qualification"], qualification)

    # ─────────────────────────────────────────────────────────
    # Запись оценок для одного предмета (isolated items)
    # ─────────────────────────────────────────────────────────

    def _write_grades(
        self,
        ws,
        row: int,
        cols: dict,
        entry: dict,
        is_kz: bool,
    ):
        """Записывает оценки одного предмета по координатам."""
        subj_obj = entry.get("subject")
        is_elective = subj_obj.is_elective if subj_obj else False
        is_practice = entry.get("is_practice", False)

        trad = entry.get("traditional_kz" if is_kz else "traditional_ru", "")
        pts = entry.get("points", "")
        let = entry.get("letter", "")
        gpa = entry.get("gpa", "")

        # Факультативы → зачтено / сынақ
        if is_elective:
            trad = self.terms.get("traditional_elective", "зачтено")
            pts, let, gpa = "", "", ""

        # Практика без оценки → зачтено
        elif is_practice and not trad:
            hours = entry.get("hours", "")
            credits_val = entry.get("credits", "")
            if not hours and not credits_val:
                trad = self.terms.get("traditional_practice", "зачтено")

        if pts:
            self._write(ws, row, cols["score_100"], pts)
        if let:
            self._write(ws, row, cols["grade_letter"], let)
        if gpa:
            self._write(ws, row, cols["gpa_numeric"], gpa)
        if trad:
            self._write(ws, row, cols["final_grade"], trad, is_trad=True)

    # ─────────────────────────────────────────────────────────
    # Merged block items 61-65 (RU only)
    # Лист 2, правая часть, row 9 (merged O9:O11, P9:P11, Q9:Q11, R9:R11)
    # ─────────────────────────────────────────────────────────

    def _fill_block_ru(self, entries: list):
        """
        Заполняет merged-блок на Лист 2 для RU (items 61-65).

        item 61 (Practice): hours/credits в M9/N9 (уже в шаблоне),
                            score/letter/gpa + traditional
        item 62 (Attestation): score/letter/gpa + traditional
        items 63-65 (Electives): only traditional
        """
        ws = self.workbook.worksheets[1]
        block_row = 9  # Merged cells start at row 9

        # R9:R11 — traditional grades (9 строк с \n)
        grade_lines = [""] * _BLOCK_TOTAL_LINES_IT

        # Practice and Attestation data for O/P/Q merged cells
        practice_score = ""
        practice_letter = ""
        practice_gpa = ""
        attest_score = ""
        attest_letter = ""
        attest_gpa = ""

        for i, entry in enumerate(entries):
            item_num = 61 + i
            if item_num > 65:
                break

            subj_obj = entry.get("subject")
            is_elective = subj_obj.is_elective if subj_obj else False

            trad = entry.get("traditional_ru", "")
            pts = entry.get("points", "")
            let = entry.get("letter", "")
            gpa = entry.get("gpa", "")

            if item_num == 61:
                # Practice: score/letter/gpa + hours/credits already in template
                practice_score = pts
                practice_letter = let
                practice_gpa = gpa
                # Hours/credits are pre-filled in template (M9=504, N9=21)
                # But write them if available from data
                hours = entry.get("hours", "")
                credits_val = entry.get("credits", "")
                if hours:
                    self._write(ws, block_row, 13, hours)  # M9
                if credits_val:
                    self._write(ws, block_row, 14, credits_val)  # N9

            elif item_num == 62:
                # Attestation: score/letter/gpa
                attest_score = pts
                attest_letter = let
                attest_gpa = gpa

            elif is_elective:
                # Electives → зачтено
                trad = self.terms.get("traditional_elective", "зачтено")

            # Traditional grade → line in R block
            line_idx = _BLOCK_LINE_INDICES_IT_RU.get(item_num)
            if line_idx is not None and trad:
                grade_lines[line_idx] = str(trad)

        # Write merged cells
        # O9:O11 — score (practice + attestation via \n)
        # Keep practice line near the top and shift only attestation by 4 Alt+Enter.
        attestation_gap = "\n" * 5  # 1 regular line break + 4 extra breaks
        if practice_score and attest_score:
            score_value = f"{practice_score}{attestation_gap}{attest_score}"
        elif practice_score:
            score_value = practice_score
        elif attest_score:
            score_value = ("\n" * 4) + attest_score
        else:
            score_value = ""
        if score_value:
            cell = ws.cell(row=block_row, column=15)
            cell.value = score_value

        # P9:P11 — letter
        if practice_letter and attest_letter:
            letter_value = f"{practice_letter}{attestation_gap}{attest_letter}"
        elif practice_letter:
            letter_value = practice_letter
        elif attest_letter:
            letter_value = ("\n" * 4) + attest_letter
        else:
            letter_value = ""
        if letter_value:
            cell = ws.cell(row=block_row, column=16)
            cell.value = letter_value

        # Q9:Q11 — gpa
        if practice_gpa and attest_gpa:
            gpa_value = f"{practice_gpa}{attestation_gap}{attest_gpa}"
        elif practice_gpa:
            gpa_value = practice_gpa
        elif attest_gpa:
            gpa_value = ("\n" * 4) + attest_gpa
        else:
            gpa_value = ""
        if gpa_value:
            cell = ws.cell(row=block_row, column=17)
            cell.value = gpa_value

        # R9:R11 — traditional grades (\n-separated block)
        cell = ws.cell(row=block_row, column=18)
        cell.value = "\n".join(grade_lines)

    # ─────────────────────────────────────────────────────────
    # Утилита записи
    # ─────────────────────────────────────────────────────────

    def _write(self, ws, row: int, col: int, val, is_trad: bool = False):
        """Записывает значение в ячейку, сохраняя alignment шаблона.

        Alignment НЕ перезаписывается — используется тот, что настроен
        в шаблоне для каждой ячейки. Это гарантирует pixel-perfect вывод.
        
        """
        cell = ws.cell(row=row, column=col)
        cell.value = val
