from __future__ import annotations

import copy
from typing import Any, Dict

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# Text anchor presets inside the start cell:
# - half   : start from middle of the cell
# - 3/4    : start from third quarter of the cell
# - end    : start near the end of the cell
# - start  : start from the beginning of the cell
INDENT_PRESETS = {
    "start": 0,
    "half": 15,
    "three_quarters": 25,
    "end": 0,
}


PAGE1_HEADER_ALIGNMENT: Dict[str, Dict[str, Dict[str, Any]]] = {
    "ru": {
        "diplom_id": {
            "row": 2,
            "col": 3,
            "horizontal": "center",
            "vertical": "center",
            "wrap_text": True,
            "indent": INDENT_PRESETS["start"],
            "zone_end_col": 4,
        },
        "full_name": {
            "row": 3,
            "col": 2,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": True,
            "indent": INDENT_PRESETS["half"],
            "zone_end_col": 5,
        },
        "start_year": {
            "row": 4,
            "col": 2,
            "horizontal": "right",
            "vertical": "center",
            "wrap_text": False,
            "indent": INDENT_PRESETS["start"],
            "zone_end_col": 3,
        },
        "end_year": {
            "row": 4,
            "col": 7,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": False,
            "indent": INDENT_PRESETS["start"],
            "zone_end_col": 7,
        },
        "college": {
            "row": 5,
            "col": 2,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": True,
            "indent": INDENT_PRESETS["half"],
            "zone_end_col": 7,
        },
        "speciality": {
            "row": 6,
            "col": 2,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": True,
            "indent": 19,
            "zone_end_col": 8,
        },
        "qualification": {
            "row": 9,
            "col": 2,
            "horizontal": "center",
            "vertical": "center",
            "wrap_text": True,
            "indent": 0,
            "zone_end_col": 8,
        },
    },
    "kz": {
        "diplom_id": {
            "row": 2,
            "col": 3,
            "horizontal": "center",
            "vertical": "center",
            "wrap_text": True,
            "indent": INDENT_PRESETS["start"],
            "zone_end_col": 4,
        },
        "full_name": {
            "row": 3,
            "col": 2,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": True,
            "indent": INDENT_PRESETS["half"],
            "zone_end_col": 5,
        },
        "start_year": {
            "row": 4,
            "col": 2,            # Колонка B
            "horizontal": "center", # В узкой ячейке лучше центрировать, чем мучиться с пробелами
            "vertical": "center",
            "wrap_text": False,
            "indent": 0,         # ОБЯЗАТЕЛЬНО 0, чтобы текст не убежал в соседнюю ячейку
            "zone_end_col": 2,   # Конечная колонка тоже B (диапазон B:B)
        },
        "end_year": {
            "row": 4,
            "col": 6,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": False,
            "indent": INDENT_PRESETS["start"],
            "zone_end_col": 6,
        },
        "college": {
            "row": 5,
            "col": 2,
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": True,
            "indent": INDENT_PRESETS["half"],
            "zone_end_col": 6,
        },
        "speciality": {
            "row": 6,
            "col": 2,
            "horizontal": "center", # Центрируем
            "vertical": "center",
            "wrap_text": True,      # Лучше вернуть True, чтобы длинный текст не обрезался
            "indent": INDENT_PRESETS["start"], # ФАКТИЧЕСКИЙ ФИКС: отступ 0
            "zone_end_col": 5,
        },
        "qualification": {
            "row": 9,
            "col": 2,
            "horizontal": "center", # Центрируем
            "vertical": "center",
            "wrap_text": True,      # Лучше вернуть True
            "indent": INDENT_PRESETS["start"], # ФАКТИЧЕСКИЙ ФИКС: отступ 0
            "zone_end_col": 5,
        },
    },
}


def get_page1_header_alignment(is_kz: bool) -> Dict[str, Dict[str, Any]]:
    return PAGE1_HEADER_ALIGNMENT["kz" if is_kz else "ru"]


def write_aligned_field(ws, field_cfg: Dict[str, Any], value: Any) -> None:
    if value is None:
        return

    text = str(value)
    if not text:
        return

    row = field_cfg["row"]
    start_col = field_cfg["col"]
    end_col = field_cfg.get("zone_end_col", start_col)

    if end_col < start_col:
        start_col, end_col = end_col, start_col

    # Keep merge boundaries in sync with alignment config.
    if start_col != end_col:
        cell_range = (
            f"{get_column_letter(start_col)}{row}:"
            f"{get_column_letter(end_col)}{row}"
        )
        try:
            ws.merge_cells(cell_range)
        except ValueError:
            # Range is already merged or overlaps an existing merged range.
            pass

    cell = ws.cell(row=row, column=start_col)
    existing_font = cell.font  # Сохраняем текущий шрифт из шаблона
    cell.value = text
    cell.font = copy.copy(existing_font)  # Восстанавливаем шрифт (цвет, bold, italic и т.д.)
    cell.alignment = Alignment(
        horizontal=field_cfg.get("horizontal", "left"),
        vertical=field_cfg.get("vertical", "center"),
        wrap_text=field_cfg.get("wrap_text", False),
        indent=field_cfg.get("indent", 0),
    )
