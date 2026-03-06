import os
import openpyxl
from openpyxl.styles import Alignment
import math
import re
import gc
from typing import Dict, Any, List

class DiplomaGenerator:
    """Единый класс для генерации Excel-дипломов на основе шаблона (openpyxl)."""
    
    def __init__(self, template_path: str, output_path: str, config: Dict[str, Any], terms: Dict[str, str]):
        self.template_path = template_path
        self.output_path = output_path
        self.config = config
        self.terms = terms
        
        # Загружаем существующий шаблон
        self.workbook = openpyxl.load_workbook(template_path)

    def is_module_header(self, subj_name: str) -> bool:
        """Определяет, является ли предмет заголовком модуля."""
        if not subj_name:
            return False
        s = str(subj_name).strip()
        return s.startswith("КМ") or s.startswith("БМ") or s.startswith("ПМ") or \
               s.startswith("Кәсіптік модуль") or s.startswith("Профессиональная практика") or \
               s.startswith("Қорытынды аттестаттау") or s.startswith("Итоговая аттестация") or \
               s.startswith("Базовые модули") or s.startswith("Профессиональные модули") or \
               s.startswith("Базалық модул") or s.startswith("Кәсіби модул")

    def is_practice(self, subj_name: str) -> bool:
        """Определяет, является ли предмет практикой (оқу, кәсіптік, профессиональная и т.д)."""
        if not subj_name: return False
        s = str(subj_name).lower()
        
        # Если в таблице есть оценка или часы, это обычная практика с оценкой
        # Но если нам нужно проставить "сынақ"/"зачтено" ТОЛЬКО для чистых практик без оценки,
        # Мы должны быть осторожны: проф. практика (Кәсіптік практика КМ3. ОН3.2...) получает оценку, а не сынақ.
        # В ТЗ: "сынак вставлен в проф практику а хотя оно не факультатив и там будет оценка"
        # Поэтому мы убираем "кәсіптік практика" и "профессиональная практика" из списка дефолтных "сынақ".
        
        practices = [
            "оқу практика", "учебная практика",
            "өндірістік практика", "производственная практика", 
            "преддипломная практика", "тәжірибе" 
        ]
        
        return any(p in s for p in practices) and "практикалық" not in s

    def is_elective(self, subj_name: str) -> bool:
        """Определяет, является ли предмет факультативом."""
        if not subj_name: return False
        return "факультатив" in str(subj_name).lower()

    def fill_student_data(self, student: Dict[str, Any]):
        """Заполняет диплом конкретными данными, сохраняя все стили и разметку шаблона."""
        grades = student['grades']
        meta = student.get('meta', {})
        from .utils import normalize_key
        
        # ──── 0. Заполняем «шапку» на первой странице ────
        ws1 = self.workbook.worksheets[0]
        self._fill_header(ws1, student, meta)
        
        # (Подсчет суммы удален, так как часы берутся напрямую из Excel для подмодулей)
                
        # ──── 2. Сквозная нумерация + заполнение оценок ────
        global_num = 0  # сквозной счётчик предметов
        
        for sheet_idx, ws in enumerate(self.workbook.worksheets):
            start_row = 15 if sheet_idx == 0 else 1
            for row in range(start_row, ws.max_row + 1):
                cell_b = ws.cell(row=row, column=2)
                subj = cell_b.value
                
                if subj and isinstance(subj, str) and subj.strip():
                    subj = subj.strip()
                    
                    global_num += 1
                    ws.cell(row=row, column=1).value = global_num
                    
                    nkey = normalize_key(subj)
                    grade = grades.get(nkey)
                    if not grade and re.match(r'((?:БМ|КМ|ПМ|ОН|РО)\s*\.?\s*\d+(?:\.\d+)?)', subj, re.IGNORECASE):
                        prefix = normalize_key(re.match(r'((?:БМ|КМ|ПМ|ОН|РО)\s*\.?\s*\d+(?:\.\d+)?)', subj, re.IGNORECASE).group(1))
                        for gk, gv in grades.items():
                            if gk.startswith(prefix):
                                grade = gv
                                break
                                
                        # Fallback for cross-language module prefixes
                        if not grade:
                            alt_prefix = prefix
                            if prefix.startswith('ро'): alt_prefix = 'он' + prefix[2:]
                            elif prefix.startswith('он'): alt_prefix = 'ро' + prefix[2:]
                            elif prefix.startswith('пм'): alt_prefix = 'км' + prefix[2:]
                            elif prefix.startswith('км'): alt_prefix = 'пм' + prefix[2:]
                            
                            if alt_prefix != prefix:
                                for gk, gv in grades.items():
                                    if gk.startswith(alt_prefix):
                                        grade = gv
                                        break

                    # Fallback для практик: "Кәсіптік практика..." / "Профессиональная практика..."
                    # Эти строки не начинаются с КМ/ПМ/ОН/РО, поэтому не попадают в prefix-lookup.
                    # Ищем по ключевым словам в grades (по subject_kz / subject_ru).
                    if not grade:
                        subj_lower = subj.lower()
                        practice_markers = ['кәсіптік практика', 'профессиональная практика']
                        if any(pm in subj_lower for pm in practice_markers):
                            for gk, gv in grades.items():
                                gv_kz = str(gv.get('subject_kz', '')).lower()
                                gv_ru = str(gv.get('subject_ru', '')).lower()
                                if any(pm in gv_kz or pm in gv_ru for pm in practice_markers):
                                    grade = gv
                                    break

                    is_header = self.is_module_header(subj)
                    hours = grade.get('hours', '') if grade else ''
                    credits_val = grade.get('credits', '') if grade else ''

                    # Агрегация часов для заголовка модуля (только для КМ и ПМ, НЕ для БМ)
                    if is_header and not hours:
                        m = re.search(r'(КМ|ПМ|БМ|ОН)\s*0*(\d+)', subj, re.IGNORECASE)
                        if m:
                            mod_type = m.group(1).lower()
                            if mod_type != 'бм':  # "Это касается всех модулей кроме базовых"
                                mod_num = m.group(2)
                                prefix_search = f"он{mod_num}" if mod_type in ('км', 'пм') else f"{mod_type}{mod_num}"
                                
                                th, tc = 0.0, 0.0
                                for gk, gv in grades.items():
                                    if gk.startswith(prefix_search):
                                        h_str = str(gv.get('hours', '0')).replace(',', '.')
                                        c_str = str(gv.get('credits', '0')).replace(',', '.')
                                        try: th += float(h_str) if h_str.replace('.', '', 1).isdigit() else 0
                                        except: pass
                                        try: tc += float(c_str) if c_str.replace('.', '', 1).isdigit() else 0
                                        except: pass
                                
                                if th > 0: hours = str(int(th)) if th.is_integer() else str(th)
                                if tc > 0: credits_val = str(int(tc)) if tc.is_integer() else str(tc)

                    # Определяем тип предмета: факультатив или практика
                    subj_kz = str(grade.get('subject_kz', '')) if grade else ''
                    is_elec = self.is_elective(subj) or self.is_elective(subj_kz)
                    is_prac = self.is_practice(subj) or self.is_practice(subj_kz)
                    
                    # Записываем часы и кредиты
                    if hours:
                        self._write_val(ws, row, 3, hours)
                        
                    if credits_val:
                        self._write_val(ws, row, 4, credits_val)
                    
                    # Пишем баллы и оценки (только если это не заголовок модуля)
                    if not is_header:
                        trad_val, pts, let, gpa = "", "", "", ""
                        
                        if grade:
                            trad_val = grade.get('traditional_kz', grade.get('traditional_ru', ''))
                            pts = grade.get('points', '')
                            let = grade.get('letter', '')
                            gpa = grade.get('gpa', '')
                            
                        # Принудительно ставим "сынақ" / "зачтено" для факультативов
                        if is_elec:
                            trad_val = self.terms.get('traditional_elective', 'зачтено')
                        elif is_prac and not trad_val:
                            # Ставим зачтено только если оценки нет И у предмета нет часов/кредитов.
                            # Если часы и кредиты есть, это полноценная дисциплина, требующая оценки.
                            if not hours and not credits_val:
                                trad_val = self.terms.get('traditional_practice', 'зачтено')
                            
                        if pts: self._write_val(ws, row, 5, pts)
                        if let: self._write_val(ws, row, 6, let)
                        if gpa: self._write_val(ws, row, 7, gpa)
                        if trad_val: self._write_val(ws, row, 8, trad_val, is_trad=True)

        print(f"Filled data for: {student['name'].encode('cp1251', errors='ignore').decode('cp1251')}")

    def _fill_header(self, ws, student: Dict[str, Any], meta: Dict[str, str]):
        """Заполняет шапку первой страницы данными студента."""
        is_kz = ws.title.lower().startswith('бет')
        
        INDENT_HALF = 4    
        INDENT_3_4 = 6     
        INDENT_END = 7     
        
        diploma_num = student.get('diploma_kz' if is_kz else 'diploma_ru', '')
        if not diploma_num or str(diploma_num) in ('nan', ''):
            diploma_num = ''
        else:
            diploma_num = re.sub(r'[^\d]', '', str(diploma_num))
        ws.cell(row=2, column=3).value = diploma_num  
        ws.cell(row=2, column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.cell(row=3, column=2).value = student['name']  
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=INDENT_HALF)
        
        year_start = meta.get('year_start', '')
        year_end = meta.get('year_end', '')
        
        if is_kz:
            if year_start:
                ws.cell(row=4, column=2).value = year_start  
                ws.cell(row=4, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=INDENT_HALF)
            if year_end:
                ws.cell(row=4, column=6).value = year_end  
                ws.cell(row=4, column=6).alignment = Alignment(horizontal='left', vertical='center')
        else:
            if year_start:
                ws.cell(row=4, column=2).value = year_start  
                ws.cell(row=4, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=INDENT_END)
            if year_end:
                ws.cell(row=4, column=6).value = year_end  
                ws.cell(row=4, column=6).alignment = Alignment(horizontal='left', vertical='center')
        
        from config.settings import INSTITUTION_NAME_KZ, INSTITUTION_NAME_RU
        if is_kz:
            college_text = INSTITUTION_NAME_KZ
        else:
            college_text = INSTITUTION_NAME_RU
        
        ws.cell(row=5, column=2).value = college_text  
        ws.cell(row=5, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=INDENT_HALF)
        
        specialty = meta.get('specialty_kz' if is_kz else 'specialty_ru', '')
        if specialty:
            ws.cell(row=6, column=2).value = specialty  
            ws.cell(row=6, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=INDENT_3_4)
        
        qualification = meta.get('qualification_kz' if is_kz else 'qualification_ru', '')
        if qualification:
            ws.cell(row=9, column=2).value = qualification  
            ws.cell(row=9, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=INDENT_3_4)

    def _write_val(self, ws, row, col, val, is_trad=False):
        """Записывает значение в ячейку, сохраняя стили шаблона."""
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.alignment = Alignment(horizontal='left' if is_trad else 'center', vertical='center', wrap_text=True)

    def close(self):
        """Сохраняет файл и очищает память."""
        self.workbook.save(self.output_path)
        self.workbook.close()
        self.workbook = None
        gc.collect()
