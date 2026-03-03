import os
import io
import openpyxl
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
import math
import time
import shutil

templates = [
    'templates/diploma_v4 (1).xlsx',
    'templates/diploma_ru_template.xlsx',
    'templates/Diplom_IT_KZ_Template.xlsx',
    'templates/Diplom_IT_RU_Template.xlsx'
]

# A4 landscape dimensions in cm
# Width: 29.7cm, Height: 21.0cm
# With 0.5 inch (1.27cm) margins top+bottom:
# Printable height: 21.0 - 2 * 1.27 = 18.46cm ≈ 18.5cm

MARGIN_INCHES = 0.5  # 1.27 cm per side
A4_LANDSCAPE_HEIGHT_CM = 21.0
PRINTABLE_HEIGHT_CM = A4_LANDSCAPE_HEIGHT_CM - 2 * MARGIN_INCHES * 2.54  # ~18.46cm

def calc_row_height(text: str, chars_per_row: int = 27) -> float:
    if not text:
        return 12.0
    text_str = str(text).strip()
    chars = len(text_str)
    explicit_lines = text_str.count('\n') + 1
    wrapped_lines = math.ceil(chars / chars_per_row)
    total_lines = max(explicit_lines, wrapped_lines)
    return max(12.0, total_lines * 9.5 + 1.0)


for t in templates:
    if not os.path.exists(t):
        print(f"SKIP: {t}")
        continue
    print(f"\nProcessing: {t}")

    with open(t, 'rb') as f:
        buf = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(buf)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # ── 1. Set landscape orientation and A4 paper ──
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9  # A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1

        # ── 2. Set margins (0.5 inches, consistent with form) ──
        ws.page_margins.left   = MARGIN_INCHES
        ws.page_margins.right  = MARGIN_INCHES
        ws.page_margins.top    = MARGIN_INCHES
        ws.page_margins.bottom = MARGIN_INCHES

        # ── 3. Recalculate row heights ──
        col_b_width = ws.column_dimensions['B'].width or 24.07
        cpl = max(5, int(col_b_width * 1.15))

        total_pts = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                h = calc_row_height(cell.value.strip(), cpl)
            else:
                h = 12.0
            ws.row_dimensions[row].height = h
            total_pts += h

        total_cm = total_pts / 72 * 2.54
        status = "✅ fits" if total_cm <= PRINTABLE_HEIGHT_CM else f"⚠️ {total_cm - PRINTABLE_HEIGHT_CM:.1f}cm over"
        print(f"  [{ws_name}] content={total_cm:.1f}cm / printable={PRINTABLE_HEIGHT_CM:.1f}cm → {status}")

    # ── Save with retry ──
    tmp_path = t + ".tmp"
    wb.save(tmp_path)

    for attempt in range(6):
        try:
            shutil.move(tmp_path, t)
            print(f"  Saved ✅")
            break
        except (PermissionError, OSError):
            if attempt < 5:
                time.sleep(2)
            else:
                print(f"  FAILED: file still locked after 6 attempts")
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

print("\n\nAll done!")
