#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Test Template Generation — Quick Smoke Test
============================================
Generates 4 test diplomas (F_KZ, F_RU, D_KZ, D_RU) using mock student data
to verify that _fill_header() correctly fills all fields in the new layout.
"""

import os
import sys
import openpyxl
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.generator import DiplomaGenerator

# ─── Mock student data ───
MOCK_STUDENT = {
    'name': 'Иванов Алексей Петрович',
    'diploma_kz': '№ ЖБ 0231383',   # should become just "0231383"
    'diploma_ru': '№ ЖБ 0231383',
    'grades': {
        'қазақтілі': {
            'hours': '72', 'credits': '3', 'points': '85',
            'letter': 'B+', 'gpa': '3.33',
            'traditional_kz': '4 (жақсы)', 'traditional_ru': '4 (хорошо)'
        },
        'казахскийязык': {
            'hours': '72', 'credits': '3', 'points': '85',
            'letter': 'B+', 'gpa': '3.33',
            'traditional_kz': '4 (жақсы)', 'traditional_ru': '4 (хорошо)'
        },
    },
    'meta': {
        'year_start': '2022',
        'year_end': '2025',
        'specialty_kz': '06130100 – Ақпараттық жүйелер (салалар бойынша)',
        'specialty_ru': '06130100 – Информационные системы (по отраслям)',
        'qualification_kz': '4S06130103 – Бағдарламашы',
        'qualification_ru': '4S06130103 – Программист',
    }
}

MOCK_TERMS_KZ = {
    'traditional_elective': 'сынақ',
    'traditional_practice': 'сынақ',
}
MOCK_TERMS_RU = {
    'traditional_elective': 'зачтено',
    'traditional_practice': 'зачтено',
}

# ─── Templates to test ───
TEMPLATES = {
    'F_KZ': ('templates/Diplom_F_KZ_Template (4).xlsx', MOCK_TERMS_KZ),
    'F_RU': ('templates/Diplom_F_RU_Template (4).xlsx', MOCK_TERMS_RU),
    'D_KZ': ('templates/Diplom_D_KZ_Template(4).xlsx', MOCK_TERMS_KZ),
    'D_RU': ('templates/Diplom_D_RU_Template(4).xlsx', MOCK_TERMS_RU),
}

OUTPUT_DIR = Path("test_output_v2")
OUTPUT_DIR.mkdir(exist_ok=True)


def inspect_output(filepath: str):
    """Inspect the generated diploma header to verify correctness."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    print(f"\n  --- Inspection of {os.path.basename(filepath)} ---")
    print(f"  Sheet: {ws.title}")
    
    # Check all header rows
    fields_to_check = {
        2: "DiplomID (C2)",
        3: "Full Name (B3)",
        4: "Dates (B4 + F4)",
        5: "College (B5)",
        6: "Specialty (B6)",
        9: "Qualification (B9)",
    }
    
    for row, label in fields_to_check.items():
        h = ws.row_dimensions[row].height
        vals = {}
        for col in range(1, 9):
            v = ws.cell(row, col).value
            a = ws.cell(row, col).alignment
            if v is not None:
                indent_str = f" indent={a.indent}" if a and a.indent else ""
                vals[chr(64+col)] = f"{v}{indent_str}"
        
        vals_str = " | ".join(f"{k}={v}" for k, v in vals.items()) if vals else "(empty)"
        print(f"  Row {row:2d} [{label:20s}] h={h:>6.1f}pt | {vals_str}")
    
    # Verify specific values
    errors = []
    
    # DiplomID should be digits only in C2
    diploma_val = ws.cell(2, 3).value
    if diploma_val and not str(diploma_val).isdigit():
        errors.append(f"DiplomID in C2 should be digits only, got: {diploma_val}")
    elif diploma_val == '0231383':
        print(f"  ✅ DiplomID correctly stripped to digits: {diploma_val}")
    
    # Full name should be in B3
    name_val = ws.cell(3, 2).value
    if name_val == 'Иванов Алексей Петрович':
        print(f"  ✅ Full name in B3: {name_val}")
    elif name_val:
        errors.append(f"Full name in wrong cell or value: {name_val}")
    
    # College text should be default
    college_val = ws.cell(5, 2).value
    is_kz = ws.title.lower().startswith('бет')
    expected_college = "Жамбыл инновациялық жоғары колледжінде" if is_kz else "Жамбылском инновационным высшем колледже"
    if college_val == expected_college:
        print(f"  ✅ College default text: {college_val[:40]}...")
    elif college_val:
        errors.append(f"College text unexpected: {college_val}")
    
    # Dates
    year_start = ws.cell(4, 2).value
    year_end = ws.cell(4, 6).value
    if year_start == '2022' and year_end == '2025':
        print(f"  ✅ Dates: start={year_start}, end={year_end}")
    
    # Row heights for rows 1-14
    expected_heights = {
        1: 24.09, 2: 21.26, 3: 28.35, 4: 21.26, 5: 21.26,
        6: 21.26, 7: 21.26, 8: 21.26, 9: 21.26, 10: 24.09,
        11: 28.35, 12: 21.26, 13: 21.26, 14: 21.26,
    }
    height_ok = True
    for r, expected in expected_heights.items():
        actual = ws.row_dimensions[r].height
        if abs(actual - expected) > 0.1:
            errors.append(f"Row {r} height: expected {expected}, got {actual}")
            height_ok = False
    if height_ok:
        print(f"  ✅ All row heights correct (rows 1-14)")
    
    # Merged cells
    merged = sorted(str(m) for m in ws.merged_cells.ranges if any(
        m.min_row <= r <= m.max_row for r in [2,3,4,5,6,9]
    ))
    print(f"  Merged cells (header area): {merged}")
    
    if errors:
        for e in errors:
            print(f"  ❌ {e}")
    else:
        print(f"  ✅ ALL CHECKS PASSED")
    
    wb.close()
    return len(errors) == 0


def main():
    print("=" * 65)
    print("  Test Diploma Generation — Mock Data")
    print("=" * 65)
    
    all_ok = True
    
    for label, (template, terms) in TEMPLATES.items():
        print(f"\n{'='*65}")
        print(f"  Generating: {label}")
        print(f"{'='*65}")
        
        if not os.path.exists(template):
            print(f"  SKIP: {template} not found")
            continue
        
        out_path = str(OUTPUT_DIR / f"test_{label}.xlsx")
        
        try:
            gen = DiplomaGenerator(template, out_path, {}, terms)
            gen.fill_student_data(MOCK_STUDENT)
            gen.close()
            print(f"  Generated: {out_path}")
            
            ok = inspect_output(out_path)
            if not ok:
                all_ok = False
        except Exception as e:
            print(f"  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            all_ok = False
    
    print(f"\n\n{'='*65}")
    if all_ok:
        print("  ✅ ALL TESTS PASSED")
    else:
        print("  ❌ SOME TESTS FAILED")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()
