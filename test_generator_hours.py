import pandas as pd
import openpyxl
import os
from src.parser import parse_excel_sheet
from src.generator import DiplomaGenerator
from configs import get_config

wb_source = openpyxl.load_workbook('local_test_copy.xlsx', read_only=True, data_only=True)
ws_source = wb_source['3D-1']
data = []
for row in ws_source.iter_rows(max_row=200, max_col=200, values_only=True):
    data.append(row)

df = pd.DataFrame(data)
students = parse_excel_sheet(df, '3D-1', start_row=5)
s = students[0]

# Generate standard KZ diploma for 3D
config, terms, template_name = get_config("3D", "kz")
template_path = os.path.join("templates", template_name)
out_path = os.path.join("output", "test_diploma.xlsx")

import sys
sys.stdout.reconfigure(encoding='utf-8')

print(f"Starting generator for {s['name']}...")
generator = DiplomaGenerator(template_path, out_path, config, terms)
generator.fill_student_data(s)
generator.close()
print("Saved to", out_path)

# Now check the saved file
wb_out = openpyxl.load_workbook(out_path, data_only=True)
ws_out = wb_out.worksheets[0]
for row in range(15, 20):
    subj = ws_out.cell(row=row, column=2).value
    hours = ws_out.cell(row=row, column=3).value
    credits = ws_out.cell(row=row, column=4).value
    grade = ws_out.cell(row=row, column=8).value
    print(f"Row {row}: {subj} | H: {hours} | C: {credits} | G: {grade}")
