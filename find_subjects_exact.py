import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

def main():
    wb = openpyxl.load_workbook('local_test_copy.xlsx', read_only=True, data_only=True)
    ws = wb['3D-1']
    
    print("=== Non-empty cells in Rows 1, 2, 3 (1-indexed 2,3,4) ===")
    
    # We will just print anything that looks like a subject header in row 2 or 3
    # openpyxl is 1-indexed for rows and cols
    for row in range(2, 5):
        print(f"\n--- ROW {row} ---")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and val.strip() and val.strip().lower() != 'nan':
                text = val.strip().replace('\n', ' ')
                if len(text) > 40:
                    text = text[:37] + "..."
                col_letter = get_column_letter(col)
                print(f"Col {col} ({col_letter}): {text}")

if __name__ == "__main__":
    main()
