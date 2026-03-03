import os
import io
import shutil
import openpyxl
import math
import time

templates = [
    'templates/diploma_v4 (1).xlsx',
    'templates/diploma_ru_template.xlsx',
    'templates/Diplom_IT_KZ_Template.xlsx',
    'templates/Diplom_IT_RU_Template.xlsx'
]

def calc_row_height(text: str, chars_per_row: int = 27) -> float:
    """Compact row height: 9.5pts/line + 1pt padding, min 12pt."""
    if not text:
        return 12.0
    text_str = str(text).strip()
    chars = len(text_str)
    explicit_lines = text_str.count('\n') + 1
    wrapped_lines = math.ceil(chars / chars_per_row)
    total_lines = max(explicit_lines, wrapped_lines)
    return max(12.0, total_lines * 9.5 + 1.0)

for t in templates:
    if not os.path.exists(t):
        print(f"SKIP: {t}")
        continue
    print(f"Fixing: {t}")
    
    with open(t, 'rb') as f:
        buf = io.BytesIO(f.read())
    
    wb = openpyxl.load_workbook(buf)
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        col_b_width = ws.column_dimensions['B'].width or 24.07
        cpl = max(5, int(col_b_width * 1.15))
        
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                ws.row_dimensions[row].height = calc_row_height(cell.value.strip(), cpl)
            else:
                ws.row_dimensions[row].height = 12.0
    
    # Save to a temp file, then keep retrying to replace the original
    tmp_path = t + ".tmp"
    wb.save(tmp_path)
    
    for attempt in range(5):
        try:
            shutil.move(tmp_path, t)
            print(f"  Saved.")
            break
        except (PermissionError, OSError) as e:
            if attempt < 4:
                print(f"  Retry {attempt+1}...")
                time.sleep(2)
            else:
                print(f"  FAILED: {e}")
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

print("\nDone!")
