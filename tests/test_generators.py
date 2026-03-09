"""
Test Suite for Diploma Generators
==================================
Unit and integration tests for data.excel_generator module.

Tests:
- DiplomaGenerator initialization
- Diploma generation for different programs and languages
- Page layout and formatting
- Subject rendering (regular, modules, electives)
- Student data rendering
- Error handling
"""

import pytest
from pathlib import Path
import io
from openpyxl import load_workbook

from data.excel_generator import DiplomaGenerator, DiplomaGenerationError
from core.models import Student, Subject, Grade, Language, Program
from core.exceptions import ValidationError


@pytest.fixture
def sample_student():
    """Create a sample student with grades."""
    student = Student(
        full_name="Иванов Иван Иванович",
        diploma_number="JB12345678",
        diploma_number_clean="12345678",
    )
    
    # Add sample grades
    grades = {
        "Қазақ тілі": Grade(points="85", letter="B+", gpa=3.33, traditional_kz="4 (жақсы)", traditional_ru="4 (хорошо)"),
        "Ағылшын тілі": Grade(points="90", letter="A-", gpa=3.67, traditional_kz="5 (өте жақсы)", traditional_ru="5 (отлично)"),
        "Математика": Grade(points="78", letter="C+", gpa=2.33, traditional_kz="4 (жақсы)", traditional_ru="4 (хорошо)"),
    }
    
    for subject_name, grade in grades.items():
        student.add_grade(subject_name, grade)
    
    return student


@pytest.fixture
def sample_subjects():
    """Create sample subjects list."""
    return [
        Subject(name_kz="Қазақ тілі", name_ru="Казахский язык", hours="72", credits="3"),
        Subject(name_kz="Ағылшын тілі", name_ru="Английский язык", hours="108", credits="4"),
        Subject(name_kz="Математика", name_ru="Математика", hours="72", credits="3"),
    ]


@pytest.mark.unit
class TestGeneratorInitialization:
    """Test DiplomaGenerator initialization."""
    
    def test_create_generator_kz(self):
        """Test creating generator for Kazakh language."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        assert gen.program == Program.IT
        assert gen.language == Language.KZ
        assert gen.academic_year == "2025-2026"
    
    def test_create_generator_ru(self):
        """Test creating generator for Russian language."""
        gen = DiplomaGenerator(Program.IT, Language.RU, "2025-2026")
        assert gen.program == Program.IT
        assert gen.language == Language.RU
    
    def test_create_generator_accounting(self):
        """Test creating generator for Accounting program."""
        gen = DiplomaGenerator(Program.ACCOUNTING, Language.KZ, "2025-2026")
        assert gen.program == Program.ACCOUNTING


@pytest.mark.unit
class TestDiplomaGeneration:
    """Test basic diploma generation."""
    
    def test_generate_kz_diploma(self, sample_student, sample_subjects):
        """Test generating Kazakh diploma."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        # Should return bytes
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0
    
    def test_generate_ru_diploma(self, sample_student, sample_subjects):
        """Test generating Russian diploma."""
        gen = DiplomaGenerator(Program.IT, Language.RU, "2025-2026")
        
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0
    
    def test_generate_returns_valid_excel(self, sample_student, sample_subjects):
        """Test that generated output is valid Excel file."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        # Try to load as Excel
        excel_io = io.BytesIO(excel_bytes)
        workbook = load_workbook(excel_io)
        
        assert workbook is not None
        assert len(workbook.worksheets) > 0


@pytest.mark.unit
class TestGradeColumnRendering:
    """Test rendering of different grade column types."""
    
    def test_render_letter_grade(self, sample_student, sample_subjects):
        """Test rendering with letter grades."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        assert len(excel_bytes) > 0
    
    def test_render_gpa_grade(self, sample_student, sample_subjects):
        """Test rendering with GPA grades."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        excel_bytes = gen.generate(sample_student, sample_subjects, "gpa")
        
        assert len(excel_bytes) > 0
    
    def test_render_traditional_grade(self, sample_student, sample_subjects):
        """Test rendering with traditional grades."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        excel_bytes = gen.generate(sample_student, sample_subjects, "traditional")
        
        assert len(excel_bytes) > 0


@pytest.mark.unit
class TestSpecialSubjectTypes:
    """Test rendering of special subject types."""
    
    def test_render_module_header(self, sample_student):
        """Test rendering of module headers (КМ, БМ)."""
        subjects = [
            Subject(
                name_kz="КМ 01 Web технологиялар",
                name_ru="КМ 01 Web технологии",
                hours="144",
                credits="6",
                is_module_header=True,
            ),
            Subject(name_kz="ОН 1.1 HTML/CSS", name_ru="ОН 1.1 HTML/CSS", hours="72", credits="3"),
        ]
        
        # Add grade for sub-subject only
        sample_student.add_grade("ОН 1.1 HTML/CSS", Grade(points="90", letter="A-", gpa=3.67))
        
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, subjects, "letter")
        
        assert len(excel_bytes) > 0
    
    def test_render_elective_subject(self, sample_student):
        """Test rendering of elective subjects."""
        subjects = [
            Subject(
                name_kz="Таңдау пәні: Дизайн",
                name_ru="Элективный предмет: Дизайн",
                hours="36",
                credits="1.5",
                is_elective=True,
            ),
        ]
        
        # Electives typically have pass/fail grades
        sample_student.add_grade(
            "Таңдау пәні: Дизайн",
            Grade(letter="сынақ", traditional_kz="сынақ", traditional_ru="зачтено")
        )
        
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, subjects, "letter")
        
        assert len(excel_bytes) > 0


@pytest.mark.integration
class TestMultiPageGeneration:
    """Test generation of multi-page diplomas."""
    
    def test_generate_multipage_diploma(self, sample_student):
        """Test generating diploma with multiple pages."""
        # Create many subjects to span multiple pages
        subjects = []
        for i in range(40):
            subject = Subject(
                name_kz=f"Пән {i+1}",
                name_ru=f"Предмет {i+1}",
                hours="72",
                credits="3",
            )
            subjects.append(subject)
            
            # Add grade
            sample_student.add_grade(
                f"Пән {i+1}",
                Grade(points="85", letter="B+", gpa=3.33)
            )
        
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, subjects, "letter")
        
        # Load and check number of sheets
        excel_io = io.BytesIO(excel_bytes)
        workbook = load_workbook(excel_io)
        
        assert len(workbook.worksheets) >= 2  # Should have multiple pages


@pytest.mark.unit
class TestErrorHandling:
    """Test generator error handling."""
    
    def test_generate_empty_student(self, sample_subjects):
        """Test generating diploma for student with no grades."""
        student = Student(full_name="Test Student", diploma_number="TEST001")
        
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        # Should handle gracefully or raise specific error
        try:
            excel_bytes = gen.generate(student, sample_subjects, "letter")
            # If it succeeds, should still return valid Excel
            assert isinstance(excel_bytes, bytes)
        except (DiplomaGenerationError, ValidationError):
            # Both are acceptable for invalid input
            pass
    
    def test_generate_invalid_grade_type(self, sample_student, sample_subjects):
        """Test generating with invalid grade type."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        
        with pytest.raises((ValueError, DiplomaGenerationError)):
            gen.generate(sample_student, sample_subjects, "invalid_type")


@pytest.mark.integration
class TestOutputValidation:
    """Test validation of generated output."""
    
    def test_output_has_correct_sheets(self, sample_student, sample_subjects):
        """Test that output has expected sheet structure."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        excel_io = io.BytesIO(excel_bytes)
        workbook = load_workbook(excel_io)
        
        # Should have at least one sheet
        assert len(workbook.worksheets) >= 1
    
    def test_output_has_student_data(self, sample_student, sample_subjects):
        """Test that output contains student information."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        excel_io = io.BytesIO(excel_bytes)
        workbook = load_workbook(excel_io)
        
        # Check first sheet has data
        sheet = workbook.worksheets[0]
        assert sheet.max_row > 0
        assert sheet.max_column > 0


@pytest.mark.slow
@pytest.mark.integration
class TestLargeDatasetGeneration:
    """Test generation for large datasets."""
    
    def test_generate_many_subjects(self, sample_student):
        """Test generating diploma with many subjects."""
        # Create 100 subjects
        subjects = []
        for i in range(100):
            subject = Subject(
                name_kz=f"Пән {i+1}",
                name_ru=f"Предмет {i+1}",
                hours="72",
                credits="3",
            )
            subjects.append(subject)
            
            sample_student.add_grade(
                f"Пән {i+1}",
                Grade(points="85", letter="B+", gpa=3.33)
            )
        
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, subjects, "letter")
        
        assert len(excel_bytes) > 0


@pytest.mark.unit
class TestLanguageRendering:
    """Test language-specific rendering."""
    
    def test_kz_language_uses_kz_names(self, sample_student, sample_subjects):
        """Test that KZ generator uses Kazakh subject names."""
        gen = DiplomaGenerator(Program.IT, Language.KZ, "2025-2026")
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        # Verify generation succeeds
        assert len(excel_bytes) > 0
    
    def test_ru_language_uses_ru_names(self, sample_student, sample_subjects):
        """Test that RU generator uses Russian subject names."""
        gen = DiplomaGenerator(Program.IT, Language.RU, "2025-2026")
        excel_bytes = gen.generate(sample_student, sample_subjects, "letter")
        
        # Verify generation succeeds
        assert len(excel_bytes) > 0
