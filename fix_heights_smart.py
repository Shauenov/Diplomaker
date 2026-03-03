import os
import io
import math
import time
import shutil
import openpyxl

templates = [
    'templates/diploma_v4 (1).xlsx',
    'templates/diploma_ru_template.xlsx',
    'templates/Diplom_IT_KZ_Template.xlsx',
    'templates/Diplom_IT_RU_Template.xlsx'
]

# A4 landscape: 210mm height. Margins 0.5in (1.27cm) top+bottom.
A4_H_CM = 21.0
MARGIN_CM = 0.5 * 2.54  # 1.27cm per side
PRINTABLE_H_CM = A4_H_CM - 2 * MARGIN_CM   # ~18.46 cm
PRINTABLE_H_PTS = PRINTABLE_H_CM / 2.54 * 72  # in points

# Minimum row height (never below this, so text is still readable)
MIN_ROW_PTS = 9.5

def calc_min_height(text: str, chars_per_row: int) -> float:
    """Minimum height needed to show all text lines."""
    if not text or not text.strip():
        return MIN_ROW_PTS
    text_str = text.strip()
    chars = len(text_str)
    explicit = text_str.count('\n') + 1
    wrapped = math.ceil(chars / chars_per_row)
    total_lines = max(explicit, wrapped)
    # line height = 9.5pt (compact but readable for 8.5pt font)
    return total_lines * 9.5


for t in templates:
    if not os.path.exists(t):
        print(f"SKIP: {t}")
        continue
    print(f"\nProcessing: {t}")

    with open(t, 'rb') as f:
        buf = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(buf)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # Set landscape, A4, fit to page
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins.left   = 0.5
        ws.page_margins.right  = 0.5
        ws.page_margins.top    = 0.5
        ws.page_margins.bottom = 0.5

        # --- Step 1: calculate minimum heights ---
        col_b_width = ws.column_dimensions['B'].width or 24.07
        # More generous chars per line (wider estimate) to reduce line count
        cpl = max(5, int(col_b_width * 1.4))  # ~33 chars

        heights = {}
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                heights[row] = calc_min_height(cell.value.strip(), cpl)
            else:
                heights[row] = MIN_ROW_PTS

        total_pts = sum(heights.values())
        total_cm = total_pts / 72 * 2.54

        # --- Step 2: Scale down if needed, respecting minimum ---
        if total_cm > PRINTABLE_H_CM:
            scale = PRINTABLE_H_PTS / total_pts
            scaled = {r: max(MIN_ROW_PTS, h * scale) for r, h in heights.items()}
            # After scaling, verify we still fit (minimums might push over)
            new_total_pts = sum(scaled.values())
            new_total_cm = new_total_pts / 72 * 2.54
            heights = scaled
            print(f"  [{ws_name}] Scaled {total_cm:.1f}cm → {new_total_cm:.1f}cm (limit {PRINTABLE_H_CM:.1f}cm)")
        else:
            print(f"  [{ws_name}] {total_cm:.1f}cm → OK (limit {PRINTABLE_H_CM:.1f}cm)")

        # --- Step 3: Apply heights ---
        for row, h in heights.items():
            ws.row_dimensions[row].height = round(h, 2)

    # Save with retry
    tmp_path = t + ".tmp"
    wb.save(tmp_path)
    for attempt in range(8):
        try:
            shutil.move(tmp_path, t)
            print(f"  Saved ✅")
            break
        except (PermissionError, OSError):
            if attempt < 7:
                time.sleep(2)
            else:
                print("  FAILED: file still locked")
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

print("\n\nAll done!")
