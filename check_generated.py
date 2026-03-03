import openpyxl

def main():
    wb = openpyxl.load_workbook('output/3D-1_Бейбутов Мухамедияр Бакытжонович_RU.xlsx', data_only=True)
    ws = wb['Лист 3']
    print("=== Sheet: Лист 3 ===")
    for row in range(1, 40):
        subj = ws.cell(row=row, column=2).value
        # if subj is string
        if subj and isinstance(subj, str) and subj.strip():
            subj_str = subj.strip()[:40].replace('\n', ' ')
            h = ws.cell(row=row, column=3).value
            c = ws.cell(row=row, column=4).value
            g = ws.cell(row=row, column=8).value
            print(f"Row {row:<2} | {subj_str:<40} | H: {str(h):<5} | C: {str(c):<5} | Grade: {str(g)}")

if __name__ == "__main__":
    main()
