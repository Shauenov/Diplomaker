import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'output\3D-1_Асанов Бекарыс Максатович_KZ.xlsx', data_only=True)
ws3 = wb['3-бет']
ws4 = wb['4-бет']

print("=== Check Submodules ===")
for row in range(1, 40):
    val = ws3.cell(row=row, column=2).value
    if val and 'ОН' in str(val):
        h = ws3.cell(row=row, column=3).value
        c = ws3.cell(row=row, column=4).value
        print(f"Row {row}: {str(val)[:30]:<30} | H: {h} C: {c}")

print("\n=== Check Electives ===")
for ws in [ws3, ws4]:
    for row in range(1, 40):
        val = ws.cell(row=row, column=2).value
        if val and 'Факультатив' in str(val):
            h = ws.cell(row=row, column=3).value
            c = ws.cell(row=row, column=4).value
            tr = ws.cell(row=row, column=8).value
            print(f"Sheet {ws.title} Row {row}: {str(val)[:30]:<30} | H: {h} C: {c} Trad: {tr}")
