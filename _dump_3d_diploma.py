"""Dump ALL entries from a generated 3D diploma to check data completeness."""
import openpyxl
import os

# Find first 3D KZ file
out_dir = "output"
files_3d = [f for f in os.listdir(out_dir) if f.startswith("3D-") and f.endswith("_KZ.xlsx") and "nan" not in f]
if not files_3d:
    print("No 3D KZ files found!")
    exit(1)

fp = os.path.join(out_dir, files_3d[0])
print(f"Checking: {files_3d[0]}")
print()

wb = openpyxl.load_workbook(fp)
for si, ws in enumerate(wb.worksheets):
    print(f"=== Sheet {si}: {ws.title} ===")
    col_shift = 1 if si in (1, 3) else 0
    
    entry_idx = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Get cell values
        col_a = row[0].value  # A column - index/entry number
        col_b = row[0 + col_shift].value if col_shift else row[0].value  # might be shifted
        
        # For shifted sheets, text is in col B (index 1)
        if col_shift:
            text_cell = row[1].value
            hours_cell = row[2].value
            credits_cell = row[3].value
            points_cell = row[4].value
            letter_cell = row[5].value
            gpa_cell = row[6].value
            trad_cell = row[7].value
        else:
            text_cell = row[1].value
            hours_cell = row[2].value
            credits_cell = row[3].value
            points_cell = row[4].value
            letter_cell = row[5].value
            gpa_cell = row[6].value
            trad_cell = row[7].value if len(row) > 7 else None
        
        if text_cell and str(text_cell).strip():
            text_short = str(text_cell).strip()[:50]
            hrs = str(hours_cell).strip() if hours_cell else "-"
            crd = str(credits_cell).strip() if credits_cell else "-"
            pts = str(points_cell).strip() if points_cell else "-"
            ltr = str(letter_cell).strip() if letter_cell else "-"
            gpa = str(gpa_cell).strip() if gpa_cell else "-"
            trd = str(trad_cell).strip()[:30] if trad_cell else "-"
            
            has_grade = pts != "-" or ltr != "-" or trd != "-"
            marker = "" if has_grade else " <<<< NO GRADE"
            
            # Check if it's БМ or ОН 4.4
            highlight = ""
            t_lower = text_short.lower()
            if "бм" in t_lower:
                highlight = " [БМ]"
            elif "он 4.4" in t_lower or "ро 4.4" in t_lower:
                highlight = " [ОН 4.4]"
            
            print(f"  {entry_idx:2d}. {text_short:50s} | h={hrs:5s} c={crd:3s} | p={pts:3s} l={ltr:3s} g={gpa:4s} t={trd}{marker}{highlight}")
            entry_idx += 1
    
    print()

wb.close()
