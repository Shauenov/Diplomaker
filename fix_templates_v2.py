#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Fix Templates v2 — Addresses all 6 issues:
1. Move subjects from row 19 → row 15 on page 1
2. Extend merged ranges to H for text fields  
3. Re-verify row heights
"""

import os, io, shutil, time
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from pathlib import Path
import copy

TEMPLATES_DIR = Path(__file__).parent / "templates"

TEMPLATES = {
    "F_KZ": TEMPLATES_DIR / "Diplom_F_KZ_Template (4).xlsx",
    "F_RU": TEMPLATES_DIR / "Diplom_F_RU_Template (4).xlsx",
    "D_KZ": TEMPLATES_DIR / "Diplom_D_KZ_Template(4).xlsx",
    "D_RU": TEMPLATES_DIR / "Diplom_D_RU_Template(4).xlsx",
}

MM_TO_PT = 72.0 / 25.4

# ─── New merged ranges (all text extends to H) ───

# Old merges to remove (from previous setup)
OLD_MERGES = {
    "KZ": ["C2:D2", "B3:E3", "B5:F5", "B6:E6", "B9:E9"],
    "RU": ["C2:D2", "B3:E3", "B4:C4", "B5:G5", "B6:H6", "B9:H9"],
}

# New merges (extend text fields to H)
NEW_MERGES = {
    "KZ": ["C2:D2", "B3:H3", "B5:H5", "B6:H6", "B9:H9"],
    "RU": ["C2:D2", "B3:H3", "B4:C4", "B5:H5", "B6:H6", "B9:H9"],
}

SUBJECT_START_OLD = 19  # Current subject start on page 1
SUBJECT_START_NEW = 15  # New subject start on page 1


def get_lang(label):
    return "KZ" if "KZ" in label else "RU"


def move_subjects_up(ws, old_start, new_start):
    """Move subjects from old_start to new_start on page 1."""
    # Collect all subject rows
    subjects = []
    for row in range(old_start, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col in range(1, 9):  # A-H
            cell = ws.cell(row, col)
            if cell.value is not None:
                has_data = True
            row_data[col] = {
                'value': cell.value,
                'font': copy.copy(cell.font),
                'alignment': copy.copy(cell.alignment),
                'border': copy.copy(cell.border),
                'fill': copy.copy(cell.fill),
                'number_format': cell.number_format,
            }
        row_data['height'] = ws.row_dimensions[row].height
        row_data['has_data'] = has_data
        subjects.append(row_data)

    # Clear old subject area (rows old_start to max_row)
    for row in range(old_start, ws.max_row + 1):
        ws.row_dimensions[row].height = 9.0
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.value = None
            cell.border = Border()

    # Also clear the buffer zone (new_start to old_start-1)
    for row in range(new_start, old_start):
        ws.row_dimensions[row].height = None  # reset to default
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.value = None

    # Write subjects at new positions
    written = 0
    for i, row_data in enumerate(subjects):
        new_row = new_start + i
        for col in range(1, 9):
            cell = ws.cell(new_row, col)
            d = row_data[col]
            cell.value = d['value']
            cell.font = d['font']
            cell.alignment = d['alignment']
            cell.border = d['border']
            cell.fill = d['fill']
            cell.number_format = d['number_format']
        if row_data['height']:
            ws.row_dimensions[new_row].height = row_data['height']
        if row_data['has_data']:
            written += 1

    return written


def fix_merged_ranges(ws, lang):
    """Remove old merges and create new ones with extended ranges."""
    existing = [str(m) for m in ws.merged_cells.ranges]

    # Remove old
    for merge_str in OLD_MERGES[lang]:
        if merge_str in existing:
            ws.unmerge_cells(merge_str)
            print(f"    Unmerged: {merge_str}")

    # Add new
    for merge_str in NEW_MERGES[lang]:
        ws.merge_cells(merge_str)
        print(f"    Merged: {merge_str}")


def process_template(label, filepath):
    print(f"\n{'='*60}")
    print(f"  {label} — {filepath.name}")
    print(f"{'='*60}")

    if not filepath.exists():
        print("  SKIP: not found")
        return

    lang = get_lang(label)

    with open(filepath, 'rb') as f:
        buf = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    # ── Step 1: Count subjects before moving ──
    subj_count = sum(1 for r in range(SUBJECT_START_OLD, ws.max_row + 1)
                     if ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip())
    print(f"\n  [1] Subjects on page 1: {subj_count} (rows {SUBJECT_START_OLD}-{ws.max_row})")

    # ── Step 2: Move subjects from row 19 → row 15 ──
    print(f"\n  [2] Moving subjects: row {SUBJECT_START_OLD} → row {SUBJECT_START_NEW}...")
    written = move_subjects_up(ws, SUBJECT_START_OLD, SUBJECT_START_NEW)
    print(f"    Moved {written} subject rows")

    # ── Step 3: Fix merged ranges (extend to H) ──
    print(f"\n  [3] Fixing merged ranges ({lang})...")
    fix_merged_ranges(ws, lang)

    # ── Step 4: Calculate subject area height ──
    total_header_mm = sum([8.5, 7.5, 10, 7.5, 7.5, 7.5, 7.5, 7.5, 7.5, 8.5, 10, 7.5, 7.5, 7.5])
    available_mm = 195 - 5  # 5mm for margins
    subject_area_mm = available_mm - total_header_mm
    print(f"\n  [4] Page 1 layout:")
    print(f"    Header area (rows 1-14): {total_header_mm} mm")
    print(f"    Available for subjects: {subject_area_mm} mm")
    print(f"    Subjects count: {subj_count}")
    if subj_count > 0:
        mm_per_subj = subject_area_mm / subj_count
        pt_per_subj = mm_per_subj * MM_TO_PT
        print(f"    Height per subject: {mm_per_subj:.2f} mm ({pt_per_subj:.1f} pt)")

        # Set subject row heights evenly
        for i in range(subj_count):
            row = SUBJECT_START_NEW + i
            ws.row_dimensions[row].height = round(pt_per_subj, 2)
        print(f"    Set {subj_count} subject rows to {pt_per_subj:.1f} pt each")

        total_used = total_header_mm + subj_count * mm_per_subj
        print(f"    Total used: {total_used:.1f} mm / {available_mm} mm available")

    # ── Step 5: Save ──
    tmp_path = str(filepath) + ".tmp"
    wb.save(tmp_path)
    wb.close()

    for attempt in range(8):
        try:
            shutil.move(tmp_path, str(filepath))
            print(f"\n  Saved!")
            break
        except (PermissionError, OSError):
            if attempt < 7:
                time.sleep(2)
            else:
                print(f"  FAILED: file locked")
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)


def verify(label, filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    print(f"\n  Verify {label}: {ws.title}")

    # Check subjects start at row 15
    first_subj_row = None
    for r in range(SUBJECT_START_NEW, ws.max_row + 1):
        if ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip():
            first_subj_row = r
            break

    if first_subj_row == SUBJECT_START_NEW:
        print(f"    Subjects start at row {SUBJECT_START_NEW}")
    else:
        print(f"    WARNING: first subject at row {first_subj_row}")

    # Count subjects
    count = sum(1 for r in range(SUBJECT_START_NEW, ws.max_row + 1)
                if ws.cell(r, 2).value and str(ws.cell(r, 2).value).strip())
    print(f"    Subject count: {count}")

    # Merged cells
    merged = sorted(str(m) for m in ws.merged_cells.ranges)
    print(f"    Merged: {merged}")

    wb.close()


def main():
    print("=" * 60)
    print("  Fix Templates v2 — Move subjects + extend merges")
    print("=" * 60)

    for label, filepath in TEMPLATES.items():
        process_template(label, filepath)

    print("\n\n" + "=" * 60)
    print("  VERIFICATION")
    print("=" * 60)

    for label, filepath in TEMPLATES.items():
        if filepath.exists():
            verify(label, filepath)

    print("\n\nDone!")


if __name__ == "__main__":
    main()
