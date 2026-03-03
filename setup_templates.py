#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Setup Templates — Row Heights & Merged Cells
=============================================
Sets precise row heights (rows 1-14) and configures merged cell ranges
for student data positioning on page 1 of all 4 diploma templates.

Physical measurements:
  - Sheet half-page: 195 mm height × 150 mm width
  - Rows 1-14: student data area = 110 mm
  - Rows 15+: subjects area = 90 mm (untouched by this script)

Column widths A-H are NOT changed (already correct).
"""

import os
import io
import shutil
import time
import openpyxl
from openpyxl.styles import Alignment
from pathlib import Path

# ─────────────────────────────────────────────────────────────
# TEMPLATES
# ─────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path(__file__).parent / "templates"

TEMPLATES = {
    "F_KZ": TEMPLATES_DIR / "Diplom_F_KZ_Template (4).xlsx",
    "F_RU": TEMPLATES_DIR / "Diplom_F_RU_Template (4).xlsx",
    "D_KZ": TEMPLATES_DIR / "Diplom_D_KZ_Template(4).xlsx",
    "D_RU": TEMPLATES_DIR / "Diplom_D_RU_Template(4).xlsx",
}

# ─────────────────────────────────────────────────────────────
# ROW HEIGHTS (mm → points: 1 mm ≈ 2.835 pt)
# ─────────────────────────────────────────────────────────────

MM_TO_PT = 72.0 / 25.4  # ≈ 2.8346

ROW_HEIGHTS_MM = {
    1:  8.5,
    2:  7.5,
    3:  10.0,
    4:  7.5,
    5:  7.5,
    6:  7.5,
    7:  7.5,
    8:  7.5,
    9:  7.5,
    10: 8.5,
    11: 10.0,
    12: 7.5,
    13: 7.5,
    14: 7.5,
}

ROW_HEIGHTS_PT = {r: round(mm * MM_TO_PT, 2) for r, mm in ROW_HEIGHTS_MM.items()}

# ─────────────────────────────────────────────────────────────
# OLD MERGED RANGES TO REMOVE (rows 1-14 area)
# ─────────────────────────────────────────────────────────────

OLD_MERGES_TO_REMOVE = [
    "A2:H2",   # old diploma ID (full width)
    "A4:H4",   # old full name (full width)
    "A6:D6",   # old start date
    "E6:H6",   # old end date
    "A7:H7",   # old college
    "A9:H9",   # old specialty (full width)
    "A12:H12", # old qualification (full width)
]

# ─────────────────────────────────────────────────────────────
# NEW MERGED RANGES PER LANGUAGE
# ─────────────────────────────────────────────────────────────

# KZ: Kazakh version
NEW_MERGES_KZ = [
    "C2:D2",   # DiplomID (numbers only)
    "B3:E3",   # Full Name
    # B4 — Start Date (single cell, no merge needed)
    # F4 — End Date (single cell, no merge needed)
    "B5:F5",   # College (default text)
    "B6:E6",   # Speciality
    "B9:E9",   # Qualification
]

# RU: Russian version
NEW_MERGES_RU = [
    "C2:D2",   # DiplomID (numbers only)
    "B3:E3",   # Full Name
    "B4:C4",   # Start Date (end of B → start of C)
    # F4 — End Date (single cell, no merge needed)
    "B5:G5",   # College (default text)
    "B6:H6",   # Speciality
    "B9:H9",   # Qualification
]


def get_lang(label: str) -> str:
    """Extract language code from template label."""
    return "KZ" if "KZ" in label else "RU"


def process_template(label: str, filepath: Path):
    """Update row heights and merged cells for page 1 of a template."""
    print(f"\n{'='*60}")
    print(f"  Processing: {label} — {filepath.name}")
    print(f"{'='*60}")

    if not filepath.exists():
        print(f"  SKIP: file not found")
        return

    lang = get_lang(label)
    new_merges = NEW_MERGES_KZ if lang == "KZ" else NEW_MERGES_RU

    # Load workbook
    with open(filepath, 'rb') as f:
        buf = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(buf)

    ws = wb.worksheets[0]  # Page 1 only
    print(f"  Sheet: {ws.title}")

    # ── Step 1: Set row heights (rows 1-14) ──
    print(f"\n  [1] Setting row heights (rows 1-14)...")
    for row, pt in ROW_HEIGHTS_PT.items():
        old_h = ws.row_dimensions[row].height
        ws.row_dimensions[row].height = pt
        print(f"      Row {row:2d}: {old_h} → {pt:.2f} pt ({ROW_HEIGHTS_MM[row]} mm)")

    # ── Step 2: Remove old merged ranges in header area ──
    print(f"\n  [2] Removing old merged ranges...")
    existing_merges = list(ws.merged_cells.ranges)
    for merge_str in OLD_MERGES_TO_REMOVE:
        found = False
        for existing in existing_merges:
            if str(existing) == merge_str:
                ws.unmerge_cells(merge_str)
                print(f"      Unmerged: {merge_str}")
                found = True
                break
        if not found:
            print(f"      Not found (skip): {merge_str}")

    # ── Step 3: Add new merged ranges ──
    print(f"\n  [3] Adding new merged ranges ({lang})...")
    for merge_str in new_merges:
        ws.merge_cells(merge_str)
        print(f"      Merged: {merge_str}")

    # ── Step 4: Save ──
    tmp_path = str(filepath) + ".tmp"
    wb.save(tmp_path)
    wb.close()

    # Retry-move to handle locked files
    for attempt in range(8):
        try:
            shutil.move(tmp_path, str(filepath))
            print(f"\n  ✅ Saved successfully")
            break
        except (PermissionError, OSError):
            if attempt < 7:
                print(f"  Retry {attempt+1}...")
                time.sleep(2)
            else:
                print(f"  ❌ FAILED: file locked")
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)


def verify_template(label: str, filepath: Path):
    """Verify that heights and merges were applied correctly."""
    print(f"\n  Verifying {label}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]

    # Check row heights
    all_ok = True
    for row, expected_pt in ROW_HEIGHTS_PT.items():
        actual = ws.row_dimensions[row].height
        ok = abs(actual - expected_pt) < 0.01
        if not ok:
            print(f"    ❌ Row {row}: expected {expected_pt:.2f}, got {actual}")
            all_ok = False

    if all_ok:
        print(f"    ✅ All row heights correct")

    # Check merged cells
    merged_strs = sorted(str(m) for m in ws.merged_cells.ranges)
    print(f"    Merged cells: {merged_strs}")

    wb.close()


def main():
    print("=" * 60)
    print("  Template Layout Setup")
    print("  Row heights (rows 1-14) + Merged cell ranges")
    print("=" * 60)

    for label, filepath in TEMPLATES.items():
        process_template(label, filepath)

    print("\n\n" + "=" * 60)
    print("  VERIFICATION")
    print("=" * 60)

    for label, filepath in TEMPLATES.items():
        if filepath.exists():
            verify_template(label, filepath)

    print("\n\nDone!")


if __name__ == "__main__":
    main()
