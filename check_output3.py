import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

def check(file_path):
    print(f"\nChecking {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print("File not found.")
        return
        
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in range(1, 65):
            val = ws.cell(row=row, column=2).value
            if val and isinstance(val, str) and any(x in val for x in ['ПМ', 'КМ', 'ОН', 'БМ', 'Ф', 'РО', 'практика', 'Практика', 'Кәсіптік', 'Профессиональная', 'Итоговая', 'Қорытынды']):
                h = ws.cell(row=row, column=3).value
                c = ws.cell(row=row, column=4).value
                pts = ws.cell(row=row, column=5).value
                tr = ws.cell(row=row, column=8).value
                print(f"Row {row:02d}: {str(val)[:40]:<40} | H: {h} C: {c} Pts: {pts} Trad: {tr}")

# IT KZ
check("output/3Ғ-1_Аймахан Балауса Абайханқызы_KZ.xlsx")
# IT RU
check("output/3Ғ-1_Аймахан Балауса Абайханқызы_RU.xlsx")
# ACC KZ
check("output/3D-1_Асанов Бекарыс Максатович_KZ.xlsx")
# ACC RU
check("output/3D-1_Асанов Бекарыс Максатович_RU.xlsx")
