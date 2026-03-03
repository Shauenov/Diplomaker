import openpyxl
import re
from src.utils import normalize_key

wb = openpyxl.load_workbook('templates/Diplom_D_RU_Template(4).xlsx', data_only=True)

# 1. First pass: find all modules and their submodules, and extract static hours from template
# Struct: { '4': { 'type': 'ПМ', 'hours': 240, 'credits': 10, 'submodules': ['ро4.1', 'ро4.2', ...] } }
modules_info = {}

for ws in wb.worksheets:
    start_row = 15 if ws.title.lower().startswith('бет') else 1
    # Actually, in Russian it doesn't start with 'бет'. Let's just do 1 to max_row
    for row in range(1, ws.max_row + 1):
        subj = ws.cell(row=row, column=2).value
        # try to parse hours
        h = ws.cell(row=row, column=3).value
        c = ws.cell(row=row, column=4).value
        
        if subj and isinstance(subj, str):
            subj_clean = subj.strip()
            # Module header
            mm = re.match(r'(ПМ|КМ|Кәсіптік|Профессиональ)[^\d]*(\d+)', subj_clean, re.IGNORECASE)
            if mm:
                m_type = mm.group(1).upper()
                m_num = mm.group(2)
                if m_num not in modules_info:
                    modules_info[m_num] = {'type': m_type, 'hours': 0, 'credits': 0, 'submodules': []}
                
                # If template has hours, save them
                if h and str(h).replace('.','',1).isdigit():
                    modules_info[m_num]['hours'] = float(str(h).replace(',','.'))
                if c and str(c).replace('.','',1).isdigit():
                    modules_info[m_num]['credits'] = float(str(c).replace(',','.'))
            
            # Submodule
            sm = re.match(r'(РО|ОН)\s*0*(\d+)\.(\d+)', subj_clean, re.IGNORECASE)
            if sm:
                p_num = sm.group(2)
                norm_name = normalize_key(subj_clean)
                if p_num not in modules_info:
                    modules_info[p_num] = {'type': 'Unknown', 'hours': 0, 'credits': 0, 'submodules': []}
                modules_info[p_num]['submodules'].append(norm_name)

print("=== Modules Found in Template ===")
for m_num, info in modules_info.items():
    print(f"Module {info['type']} {m_num}: H={info['hours']}, C={info['credits']}")
    subs = info['submodules']
    print(f"  Submodules ({len(subs)}): {subs}")
    if len(subs) > 0:
        dh = info['hours'] / len(subs) if info['hours'] else 0
        dc = info['credits'] / len(subs) if info['credits'] else 0
        print(f"  -> Calculated per sub: H={dh}, C={dc}")

