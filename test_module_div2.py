import openpyxl
import re

wb = openpyxl.load_workbook('templates/Diplom_D_RU_Template(4).xlsx', data_only=True)

modules_info = {}

for ws in wb.worksheets:
    for row in range(1, ws.max_row + 1):
        subj = ws.cell(row=row, column=2).value
        h = ws.cell(row=row, column=3).value
        c = ws.cell(row=row, column=4).value
        
        if subj and isinstance(subj, str):
            subj_clean = subj.strip()
            
            # Check if module header: "ПМ 4", "КМ 02", "Профессиональная практика 2"
            mm = re.match(r'(ПМ|КМ|БМ|Кәсіптік|Профессиональ|Оқу|Учебн)[^\d]*(\d+)', subj_clean, re.IGNORECASE)
            if mm:
                m_num = mm.group(2).lstrip('0')
                if m_num not in modules_info:
                    modules_info[m_num] = {'hours': 0, 'credits': 0, 'subrows': []}
                
                if h is not None and str(h).strip():
                    try: modules_info[m_num]['hours'] = float(str(h).replace(',','.'))
                    except: pass
                if c is not None and str(c).strip():
                    try: modules_info[m_num]['credits'] = float(str(c).replace(',','.'))
                    except: pass

            # Check if submodule: "РО 4.1", "ОН 4.1"
            sm = re.match(r'(РО|ОН)\s*0*(\d+)\.\d+', subj_clean, re.IGNORECASE)
            if sm:
                p_num = sm.group(2).lstrip('0')
                if p_num not in modules_info:
                    modules_info[p_num] = {'hours': 0, 'credits': 0, 'subrows': []}
                modules_info[p_num]['subrows'].append(row)

import sys
sys.stdout.reconfigure(encoding='utf-8')

print("=== TEMPLATE MODULES ===")
for m_num, info in modules_info.items():
    subs = info['subrows']
    print(f"Module {m_num}: H={info['hours']} C={info['credits']}, Submodules count: {len(subs)}")
    if len(subs) > 0 and info['hours'] > 0:
        print(f"  -> We should divide {info['hours']} by {len(subs)} = {info['hours']/len(subs)} hours per submodule")

