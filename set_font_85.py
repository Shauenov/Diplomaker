import os
import io
import openpyxl
from openpyxl.styles import Font

# Update all 4 templates
templates = [
    'templates/diploma_v4 (1).xlsx',
    'templates/diploma_ru_template.xlsx',
    'templates/Diplom_IT_KZ_Template.xlsx',
    'templates/Diplom_IT_RU_Template.xlsx'
]

for t in templates:
    if not os.path.exists(t):
        continue
    print(f"Applying size 8.5 font to {t} ...")
    
    with open(t, "rb") as f:
        in_mem = io.BytesIO(f.read())
        
    wb = openpyxl.load_workbook(in_mem)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Keep existing font properties, just change size to 8.5
                    if cell.font:
                        new_font = openpyxl.styles.Font(
                            name=cell.font.name or 'Times New Roman',
                            size=8.5,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                        cell.font = new_font
                    else:
                        cell.font = openpyxl.styles.Font(name='Times New Roman', size=8.5)

    wb.save(t)
    print(f"Saved {t}.")
