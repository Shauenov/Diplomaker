#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Inspect all 4 templates structure."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

templates = [
    'templates/Diplom_F_KZ_Template (4).xlsx',
    'templates/Diplom_F_RU_Template (4).xlsx',
    'templates/Diplom_D_KZ_Template(4).xlsx',
    'templates/Diplom_D_RU_Template(4).xlsx',
]

for tpath in templates:
    print("=" * 70)
    print(f"TEMPLATE: {tpath}")
    wb = openpyxl.load_workbook(tpath)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        subj_rows = []
        for row in range(1, ws.max_row + 1):
            b = ws.cell(row, 2).value
            if b and str(b).strip():
                subj_rows.append((row, str(b).strip()[:60]))
        print(f"  {ws_name}: max_row={ws.max_row}, max_col={ws.max_column}, subjects={len(subj_rows)}")
        for r, name in subj_rows:
            print(f"    R{r:2d}: {name}")
    wb.close()
    print()
